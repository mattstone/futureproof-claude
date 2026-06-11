"""
Build FutureProof 5-Year Financial Model — 3-case version.

Source-of-truth for: Conservative / Likely / Optimistic case P&Ls used in the
capital-raise pitch (see docs/capital_raise/deck_outline.md slides 9, 11, E1).

Aligned to the v14d Optimised Parameters product spec:
- FP margin: 50 bps annual (flat)
- Profit share: 10% at 3-year resets
- 50% surplus split at year-30 maturity (beyond 5-year window)
- Asymmetric hedge collar (cap +40% / floor -20%)
- PoC year-30 = 0.03 (3% portfolio capital shortfall probability)

To re-run after changing drivers:
    python3 docs/capital_raise/financial_model/build_model.py
"""

from __future__ import annotations
from dataclasses import dataclass, field
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "data/FutureProof_5Year_Model_v10.2_3Cases.xlsx"

# ---------------------------------------------------------------------------
# Drivers - the 3 cases
# Each driver is documented with the narrative reason it differs across cases.
# ---------------------------------------------------------------------------

@dataclass
class CaseDrivers:
    name: str
    narrative: str

    # --- Sales / volume ---
    licensees_uk_au: list[int]      # new licensees per year, Y1..Y5
    licensees_usa: list[int]
    mortgages_per_licensee_y1: int
    mortgages_growth_pa: float       # per year
    y1_conversion_rate: float        # fraction of expected vol in licensee's first year

    # --- Pricing (aligned to EpmModelConfig v14d Optimised) ---
    fp_annual_margin_bps: float       # FP margin charged on AUM annually (flat, not ramping)
                                       # v14d = 50 bps
    onboarding_fee_per_licensee_usd: float
    cap_markets_line_per_licensee_usd: float
    cap_markets_fee_bps: float       # gross arranger fee
    fp_commission_pct: float         # FP share of arranger fee
    ltv: float
    home_value_uk_au_usd: float
    home_value_usa_usd: float

    # --- Investment & profit share (aligned to EpmModelConfig v14d Optimised) ---
    sp500_lt_return: float            # annual total return assumed for surplus accrual
    profit_share_pct: float           # FP profit share at K-yr reset (v14d = 0.10 = 10%)
    profit_share_reset_years: int     # K = reset cadence in years (v14d = 3)

    # --- Costs ---
    fte: list[float]                 # Y0..Y5 (length 6)
    avg_loaded_salary_usd: float
    salary_growth_pa: float
    marketing_pct_of_rev: list[float]  # Y0..Y5 (length 6)
    per_fte_overhead_usd: float      # rent+IT+telecom+stationery+insurance+sundry per FTE/yr
    sales_travel_per_licensee_usd: float
    legal_fees_pct_rev: float        # legal as % of revenue
    actuarial_pct_rev: float
    audit_pct_rev: float
    platform_build_y0_usd: float
    platform_build_decay: float       # how fast platform build cost falls per year
    contingency_pct: float

    # --- AI ---
    ai_infra_y0_usd: float
    ai_infra_growth_pa: float

    # --- Other ---
    interest_rate_on_cash: float
    other_income_growth: float        # other income (captive insurer etc) growth pa


CONSERVATIVE = CaseDrivers(
    name="Conservative",
    narrative=(
        "Genuine downside scenario — ~25% probability mass at or below. "
        "Year-1 AU licensee slips into Y2 (super-fund procurement runs longer than expected); "
        "USA launch slips to late Y3 with limited cohort uptake; equity markets deliver "
        "below-LT-average (6%) over the 5-yr window. v14d Optimised EPM product spec "
        "(50 bps FP margin, 10% profit share at 3-year resets, asymmetric collar)."
    ),
    licensees_uk_au=[0, 1, 1, 2, 2],            # 6 cumulative — Y1 sales miss
    licensees_usa=[0, 0, 1, 1, 1],              # 3 cumulative — slower US launch + lower uptake
    mortgages_per_licensee_y1=2000,
    mortgages_growth_pa=0.06,
    y1_conversion_rate=0.40,
    fp_annual_margin_bps=50,                     # v14d Optimised
    onboarding_fee_per_licensee_usd=1_500_000,
    cap_markets_line_per_licensee_usd=75_000_000,
    cap_markets_fee_bps=75,
    fp_commission_pct=0.20,
    ltv=0.80,
    home_value_uk_au_usd=1_200_000,
    home_value_usa_usd=750_000,
    sp500_lt_return=0.06,                        # well below LT avg in 5-yr window
    profit_share_pct=0.10,                       # v14d Optimised
    profit_share_reset_years=3,                  # v14d Optimised — 3-yr reset cadence
    fte=[9.5, 14, 18, 22, 26, 30],
    avg_loaded_salary_usd=300_000,
    salary_growth_pa=0.05,
    marketing_pct_of_rev=[0.0, 0.12, 0.10, 0.08, 0.07, 0.05],
    per_fte_overhead_usd=45_000,
    sales_travel_per_licensee_usd=90_000,
    legal_fees_pct_rev=0.04,
    actuarial_pct_rev=0.025,
    audit_pct_rev=0.012,
    platform_build_y0_usd=3_000_000,
    platform_build_decay=0.55,
    contingency_pct=0.04,
    ai_infra_y0_usd=120_000,
    ai_infra_growth_pa=0.45,
    interest_rate_on_cash=0.04,
    other_income_growth=1.4,
)

LIKELY = CaseDrivers(
    name="Likely",
    narrative=(
        "Central tendency — ~50% probability mass within ±25% of these outcomes. "
        "Accenture channel delivers on plan: AU launch on time Q4/2026 with 1 super-fund partner; "
        "1–2 new licensees per year through Y5; USA launch Q4/2027 with steady (not explosive) ramp. "
        "Equity returns at 8% long-term — defensive vs the 10% historical headline. v14d Optimised "
        "EPM product spec: 50 bps FP margin, 10% profit share at 3-year resets, asymmetric collar "
        "(cap +40% / floor -20%); PoC year-30 = 0.03."
    ),
    licensees_uk_au=[1, 2, 2, 2, 1],            # 8 cumulative — slower Y5 cadence
    licensees_usa=[0, 0, 1, 2, 3],              # 6 cumulative — USA launch in Y3
    mortgages_per_licensee_y1=3000,
    mortgages_growth_pa=0.08,
    y1_conversion_rate=0.55,
    fp_annual_margin_bps=50,                     # v14d Optimised
    onboarding_fee_per_licensee_usd=1_500_000,
    cap_markets_line_per_licensee_usd=100_000_000,
    cap_markets_fee_bps=75,
    fp_commission_pct=0.20,
    ltv=0.80,
    home_value_uk_au_usd=1_200_000,
    home_value_usa_usd=750_000,
    sp500_lt_return=0.08,                        # defensive LT — net-of-friction realisable
    profit_share_pct=0.10,                       # v14d Optimised
    profit_share_reset_years=3,                  # v14d Optimised
    fte=[9.5, 16, 25, 33, 38, 38],
    avg_loaded_salary_usd=300_000,
    salary_growth_pa=0.05,
    marketing_pct_of_rev=[0.0, 0.10, 0.075, 0.05, 0.05, 0.025],
    per_fte_overhead_usd=45_000,
    sales_travel_per_licensee_usd=90_000,
    legal_fees_pct_rev=0.035,
    actuarial_pct_rev=0.025,
    audit_pct_rev=0.012,
    platform_build_y0_usd=3_000_000,
    platform_build_decay=0.55,
    contingency_pct=0.04,
    ai_infra_y0_usd=180_000,
    ai_infra_growth_pa=0.50,
    interest_rate_on_cash=0.04,
    other_income_growth=1.4,
)

OPTIMISTIC = CaseDrivers(
    name="Optimistic",
    narrative=(
        "Strong sales execution + favourable equity-return environment. "
        "Multi-fund RIC compliance pressure compounds in Y2-Y3 — 2-3 new licensees per year. "
        "Equity returns at 10% long-term (historical headline). v14d Optimised EPM product spec: "
        "50 bps FP margin, 10% profit share at 3-year resets, asymmetric collar (cap +40% / floor "
        "-20%); PoC year-30 = 0.03. Per-mortgage 30yr mean FP revenue ~$1.04M USD."
    ),
    licensees_uk_au=[2, 3, 3, 3, 3],            # 14 cumulative
    licensees_usa=[0, 2, 3, 4, 5],              # 14 cumulative
    mortgages_per_licensee_y1=4500,
    mortgages_growth_pa=0.12,
    y1_conversion_rate=0.75,
    fp_annual_margin_bps=50,                     # v14d Optimised
    onboarding_fee_per_licensee_usd=1_500_000,
    cap_markets_line_per_licensee_usd=150_000_000,
    cap_markets_fee_bps=75,
    fp_commission_pct=0.20,
    ltv=0.80,
    home_value_uk_au_usd=1_200_000,
    home_value_usa_usd=750_000,
    sp500_lt_return=0.10,                        # historical LT headline
    profit_share_pct=0.10,                       # v14d Optimised
    profit_share_reset_years=3,                  # v14d Optimised
    fte=[9.5, 18, 30, 40, 48, 55],
    avg_loaded_salary_usd=300_000,
    salary_growth_pa=0.05,
    marketing_pct_of_rev=[0.0, 0.08, 0.06, 0.04, 0.04, 0.02],
    per_fte_overhead_usd=45_000,
    sales_travel_per_licensee_usd=90_000,
    legal_fees_pct_rev=0.030,
    actuarial_pct_rev=0.022,
    audit_pct_rev=0.011,
    platform_build_y0_usd=3_000_000,
    platform_build_decay=0.55,
    contingency_pct=0.04,
    ai_infra_y0_usd=240_000,
    ai_infra_growth_pa=0.55,
    interest_rate_on_cash=0.04,
    other_income_growth=1.4,
)

CASES = [CONSERVATIVE, LIKELY, OPTIMISTIC]

# Constants
USD_AUD_FX = 1.50  # 1 USD = 1.50 AUD (place-holder; update with current rate before sending)
YEARS = ["Year 0", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
YEAR_INDEX = list(range(6))  # 0..5


# ---------------------------------------------------------------------------
# P&L computation
# ---------------------------------------------------------------------------

def compute_pnl(d: CaseDrivers) -> dict:
    """Compute the full P&L for one case. Returns a dict of keyed series."""

    out = {}

    # ----- Volume / book build -----
    # Mortgages per licensee per year (target volume per licensee in year N)
    mortgages_per_lic = [
        d.mortgages_per_licensee_y1 * (1 + d.mortgages_growth_pa) ** y
        for y in range(5)  # Y1..Y5
    ]

    # New mortgages written each year, by region
    # In a licensee's FIRST year, they write at y1_conversion_rate of target.
    # In subsequent years, they write at full target.
    new_mort_uk_au = [0]  # Y0
    new_mort_usa = [0]
    cum_lic_uk_au = [0]
    cum_lic_usa = [0]
    for y in range(5):  # Y1..Y5 indexed 0..4
        # licensees signed in this year (will write at y1_conversion_rate this year)
        new_lic_uk = d.licensees_uk_au[y]
        new_lic_us = d.licensees_usa[y]
        # licensees who were already on board (at full velocity)
        existing_lic_uk = cum_lic_uk_au[-1]
        existing_lic_us = cum_lic_usa[-1]

        new_uk_this_yr = (
            existing_lic_uk * mortgages_per_lic[y]
            + new_lic_uk * mortgages_per_lic[y] * d.y1_conversion_rate
        )
        new_us_this_yr = (
            existing_lic_us * mortgages_per_lic[y]
            + new_lic_us * mortgages_per_lic[y] * d.y1_conversion_rate
        )

        new_mort_uk_au.append(new_uk_this_yr)
        new_mort_usa.append(new_us_this_yr)
        cum_lic_uk_au.append(existing_lic_uk + new_lic_uk)
        cum_lic_usa.append(existing_lic_us + new_lic_us)

    cum_mort_uk_au = [sum(new_mort_uk_au[:i + 1]) for i in range(6)]
    cum_mort_usa = [sum(new_mort_usa[:i + 1]) for i in range(6)]
    cum_mort_total = [a + b for a, b in zip(cum_mort_uk_au, cum_mort_usa)]

    # Average loan size per region
    loan_size_uk_au = d.ltv * d.home_value_uk_au_usd
    loan_size_usa = d.ltv * d.home_value_usa_usd

    # AUM = book value of loans outstanding
    aum_uk_au = [m * loan_size_uk_au for m in cum_mort_uk_au]
    aum_usa = [m * loan_size_usa for m in cum_mort_usa]
    aum_total = [a + b for a, b in zip(aum_uk_au, aum_usa)]

    # ----- Revenue -----

    # Stream 1: Onboarding fee (one-off per new licensee)
    onboarding = [d.onboarding_fee_per_licensee_usd]  # Y0: founding licensee
    for y in range(5):
        onb = (d.licensees_uk_au[y] + d.licensees_usa[y]) * d.onboarding_fee_per_licensee_usd
        # Y0 also gets initial founding licensee onboarding; Y1+ fees from new sign-ups
        if y == 0:
            # Onboarding for Y1 new licensees + 2 from Y0 already in onboarding[0]
            # Match v10.1 cadence: Y1 = $4.5M -> means 3 onboardings (1 Y0 launch + 2 Y1)
            # We'll keep simple: Y0 has founder fee, Y1+ has new-signups fee
            pass
        onboarding.append(onb)

    # Stream 2: SaaS recurring — FP annual margin × cumulative book (FLAT bps per v14a)
    loan_margins_bps = [0] + [d.fp_annual_margin_bps] * 5  # Y0=0, Y1..Y5 flat
    saas = [aum * (m / 10_000) for aum, m in zip(aum_total, loan_margins_bps)]

    # Stream 3: Profit share at K-yr resets (v14d: K=3, profit_share_pct=10%)
    # 50% surplus split at year-30 maturity is separate (beyond the 5-year window).
    K = d.profit_share_reset_years
    per_mort_surplus_uk_K = (
        loan_size_uk_au * ((1 + d.sp500_lt_return) ** K - 1) * d.profit_share_pct
    )
    per_mort_surplus_us_K = (
        loan_size_usa * ((1 + d.sp500_lt_return) ** K - 1) * d.profit_share_pct
    )
    # For each vintage Y_v, the first reset hits at Y_v + K. Within the 5-year
    # window, we capture only the FIRST reset for each vintage that fits.
    surplus = [0] * 6  # Y0..Y5
    for vintage_year in range(1, 6):  # Y1..Y5
        first_reset_year = vintage_year + K
        if first_reset_year <= 5:
            cohort_surplus = (
                new_mort_uk_au[vintage_year] * per_mort_surplus_uk_K
                + new_mort_usa[vintage_year] * per_mort_surplus_us_K
            )
            surplus[first_reset_year] += cohort_surplus

    # Stream 4: Capital markets arranger (per cumulative licensee per year)
    cap_mkts = [0]  # Y0
    for y in range(5):
        cum_lic = cum_lic_uk_au[y + 1] + cum_lic_usa[y + 1]
        # New funding line per licensee per year × gross fee × FP cut
        cm = (
            cum_lic
            * d.cap_markets_line_per_licensee_usd
            * (d.cap_markets_fee_bps / 10_000)
            * d.fp_commission_pct
        )
        cap_mkts.append(cm)

    # Stream 5: Other income (interest on cash, captive insurer)
    # Simplified: $0 Y0, then growing series tied to scale
    other = [0]
    base_other = 200_000
    for y in range(5):
        other.append(base_other * (d.other_income_growth ** y))

    # Total revenue (recurring = ex-surplus)
    rev_recurring = [
        o + s + cm + ot
        for o, s, cm, ot in zip(onboarding, saas, cap_mkts, other)
    ]
    rev_total = [r + sp for r, sp in zip(rev_recurring, surplus)]

    # ----- Costs -----

    # Salaries (FTE × loaded salary × indexation)
    salaries = [
        d.fte[y] * d.avg_loaded_salary_usd * (1 + d.salary_growth_pa) ** y
        for y in range(6)
    ]

    # Per-FTE overheads
    overheads = [d.fte[y] * d.per_fte_overhead_usd * (1 + d.salary_growth_pa) ** y for y in range(6)]

    # Marketing (% of recurring revenue, with floor)
    marketing = []
    for y in range(6):
        m = max(d.marketing_pct_of_rev[y] * rev_recurring[y], 200_000 if y == 0 else 0)
        marketing.append(m)

    # Sales travel (per cumulative licensee)
    sales_travel = [0]
    for y in range(5):
        cum_lic = cum_lic_uk_au[y + 1] + cum_lic_usa[y + 1]
        sales_travel.append(cum_lic * d.sales_travel_per_licensee_usd)

    # Professional services as % of recurring revenue
    legal = [r * d.legal_fees_pct_rev for r in rev_recurring]
    actuarial = [r * d.actuarial_pct_rev for r in rev_recurring]
    audit = [r * d.audit_pct_rev for r in rev_recurring]
    # Y0 floors for professional services (regulatory work pre-revenue)
    legal[0] = max(legal[0], 700_000)
    actuarial[0] = max(actuarial[0], 300_000)
    audit[0] = max(audit[0], 80_000)

    # Platform build / IT development (front-loaded)
    platform = [d.platform_build_y0_usd]
    for y in range(5):
        platform.append(platform[-1] * d.platform_build_decay)

    # AI infrastructure (LLM API + AI eng tooling)
    ai_infra = [d.ai_infra_y0_usd * (1 + d.ai_infra_growth_pa) ** y for y in range(6)]

    # IP / patents / non-exec director fees (small, fixed-ish)
    ip_legal = [0, 350_000, 367_500, 385_875, 405_169, 425_427]
    ned_fees = [0, 100_000, 105_000, 110_250, 115_763, 121_551]

    # Bank fees (small, %-of-rev)
    bank_fees = [r * 0.001 for r in rev_recurring]
    bank_fees[0] = max(bank_fees[0], 10_000)

    total_opex = []
    for y in range(6):
        total_opex.append(
            salaries[y] + overheads[y] + marketing[y] + sales_travel[y]
            + legal[y] + actuarial[y] + audit[y]
            + platform[y] + ai_infra[y]
            + ip_legal[y] + ned_fees[y] + bank_fees[y]
        )

    contingency = [t * d.contingency_pct for t in total_opex]
    total_opex_with_cont = [t + c for t, c in zip(total_opex, contingency)]

    # ----- EBITDA -----
    ebitda = [r - c for r, c in zip(rev_total, total_opex_with_cont)]
    ebitda_recurring = [r - c for r, c in zip(rev_recurring, total_opex_with_cont)]
    ebitda_pct_total = [
        (e / r) if r > 0 else 0 for e, r in zip(ebitda, rev_total)
    ]
    ebitda_pct_recurring = [
        (e / r) if r > 0 else 0 for e, r in zip(ebitda_recurring, rev_recurring)
    ]

    # ----- Operational metrics -----
    opex_per_aum_bps = [
        (c / a * 10_000) if a > 0 else None
        for c, a in zip(total_opex_with_cont, aum_total)
    ]

    # Cash (cumulative, before any raise)
    cumulative_cash_burn = []
    running = 0
    for e in ebitda_recurring:  # use recurring for runway sanity
        running += e
        cumulative_cash_burn.append(running)

    # ----- Pack output -----
    out["mortgages_per_lic"] = [None] + mortgages_per_lic  # Y0=None, Y1..Y5
    out["licensees_uk_au_new"] = [0] + d.licensees_uk_au
    out["licensees_usa_new"] = [0] + d.licensees_usa
    out["cum_lic_uk_au"] = cum_lic_uk_au
    out["cum_lic_usa"] = cum_lic_usa
    out["new_mort_uk_au"] = new_mort_uk_au
    out["new_mort_usa"] = new_mort_usa
    out["cum_mort_uk_au"] = cum_mort_uk_au
    out["cum_mort_usa"] = cum_mort_usa
    out["cum_mort_total"] = cum_mort_total
    out["aum_uk_au"] = aum_uk_au
    out["aum_usa"] = aum_usa
    out["aum_total"] = aum_total
    out["loan_margins_bps"] = loan_margins_bps

    out["onboarding"] = onboarding
    out["saas"] = saas
    out["surplus"] = surplus
    out["cap_mkts"] = cap_mkts
    out["other"] = other
    out["rev_recurring"] = rev_recurring
    out["rev_total"] = rev_total

    out["salaries"] = salaries
    out["overheads"] = overheads
    out["marketing"] = marketing
    out["sales_travel"] = sales_travel
    out["legal"] = legal
    out["actuarial"] = actuarial
    out["audit"] = audit
    out["platform"] = platform
    out["ai_infra"] = ai_infra
    out["ip_legal"] = ip_legal
    out["ned_fees"] = ned_fees
    out["bank_fees"] = bank_fees
    out["total_opex"] = total_opex
    out["contingency"] = contingency
    out["total_opex_with_cont"] = total_opex_with_cont

    out["ebitda"] = ebitda
    out["ebitda_recurring"] = ebitda_recurring
    out["ebitda_pct_total"] = ebitda_pct_total
    out["ebitda_pct_recurring"] = ebitda_pct_recurring

    out["fte"] = d.fte
    out["opex_per_aum_bps"] = opex_per_aum_bps
    out["cumulative_cash_burn"] = cumulative_cash_burn

    return out


# ---------------------------------------------------------------------------
# Excel construction
# ---------------------------------------------------------------------------

# Styles
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
H1_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
H2_FONT = Font(name="Calibri", size=11, bold=True)
TOTAL_FONT = Font(name="Calibri", size=11, bold=True)
NORMAL_FONT = Font(name="Calibri", size=10)
NOTE_FONT = Font(name="Calibri", size=9, italic=True, color="666666")

NAVY_FILL = PatternFill("solid", fgColor="1F3864")
SECTION_FILL = PatternFill("solid", fgColor="2E5597")
LIGHT_FILL = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FILL = PatternFill("solid", fgColor="FFE699")
EBITDA_FILL = PatternFill("solid", fgColor="C6E0B4")

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

USD_FMT = "#,##0;(#,##0);-"
USD_M_FMT = '#,##0.0,,"M";(#,##0.0,,"M");-'
PCT_FMT = "0.0%"
BPS_FMT = '0.0" bps"'
CNT_FMT = "#,##0"


def style_title_row(ws, row, text, span=8):
    ws.cell(row=row, column=1, value=text).font = TITLE_FONT
    ws.cell(row=row, column=1).fill = NAVY_FILL
    ws.cell(row=row, column=1).alignment = Alignment(vertical="center", horizontal="left")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 28


def style_section_row(ws, row, text, span=8, fill=SECTION_FILL):
    c = ws.cell(row=row, column=1, value=text)
    c.font = H1_FONT
    c.fill = fill
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = 22


def write_label_year_row(ws, row, label, values, fmt=USD_FMT, font=None, fill=None,
                          start_col=2, indent=False):
    """Write a row: label at col 1, then values across columns starting at start_col."""
    c = ws.cell(row=row, column=1, value=label)
    if font:
        c.font = font
    else:
        c.font = NORMAL_FONT
    if fill:
        c.fill = fill
    if indent:
        c.alignment = Alignment(indent=1)
    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=start_col + i, value=v)
        if isinstance(v, (int, float)):
            cell.number_format = fmt
        if font:
            cell.font = font
        if fill:
            cell.fill = fill


def build_readme(wb):
    ws = wb.create_sheet("README", 0)
    ws.column_dimensions["A"].width = 110

    rows = [
        ("FutureProof — 5-Year Financial Model v10.2 (3-Case)", "title"),
        ("", ""),
        ("Purpose", "h1"),
        ("Pitch-stage 3-case financial model used in capital_raise materials. Conservative / Likely / Optimistic.", ""),
        ("", ""),
        ("Source of truth — EPM product economics (v14d Optimised)", "h1"),
        ("Pricing and surplus mechanics in this model are aligned to EpmModelConfig v14d Optimised (per the optimised parameter set).", ""),
        ("Key v14d parameters used: fp_margin = 0.50% (50 bps, all cases); profit_share_pct = 10% at 3-year resets; LTV = 80%; asymmetric hedge collar (cap +40% / floor -20%); equity stochastic mean 9.4% / vol 17.5%.", ""),
        ("Pavel Monte Carlo (50,000 paths) validated risk metrics: PoC year-30 = 0.03 (3% portfolio capital shortfall probability); per-mortgage PoD year-30 = 7.7%; mean per-mortgage surplus year-30 = $1.17M; mean per-mortgage FP revenue (full 30-yr life) = ~$1.04M.", ""),
        ("All three cases use the same v14d Optimised product spec. Differentiation between cases comes from sales velocity (licensee count, mortgage volume per licensee, conversion rate) and equity-return assumptions (Conservative 6%, Likely 8%, Optimistic 10%) — not from product-spec variation.", ""),
        ("", ""),
        ("If diligence requires v10.1 reconciliation, document the driver-by-driver delta separately — do NOT retrofit v10.2 numbers to v10.1.", ""),
        ("", ""),
        ("Tabs", "h1"),
        ("  Drivers — every load-bearing assumption, with 3 case columns and the narrative rationale per case.", ""),
        ("  P&L_Conservative / P&L_Likely / P&L_Optimistic — full 5-year P&L per case, with operational metrics.", ""),
        ("  Summary_3Case — side-by-side comparison of revenue, EBITDA, AUM, FTE, opex/AUM across all three.", ""),
        ("  AU_Only — AU-launch carve-out (UK/AU portion only) for the AU-led raise narrative.", ""),
        ("  Surplus_Sensitivity — Y5 surplus realisation under 4 different S&P long-term return assumptions (5/7/10/12%).", ""),
        ("  Runway_vs_Raise — cash runway under A$5M / A$7.5M / A$10M raise scenarios for each case.", ""),
        ("  AI_Cost_Detail — LLM API spend, AI engineering costs, AI infra as % of total opex.", ""),
        ("", ""),
        ("Currency", "h1"),
        ("All P&L figures in USD. Round size and AU market figures in AUD. FX assumption: 1 USD = 1.50 AUD (placeholder — update before pitch).", ""),
        ("Update FX at the top of the Drivers tab — flows through to AU-only and Runway tabs.", ""),
        ("", ""),
        ("Source-of-truth and reproducibility", "h1"),
        ("This file was generated by docs/capital_raise/financial_model/build_model.py. To rerun with different drivers:", ""),
        ("    python3 docs/capital_raise/financial_model/build_model.py", ""),
        ("To change driver values, edit the CaseDrivers definitions in the Python script. Re-run produces a fresh xlsx.", ""),
        ("", ""),
        ("Honest caveats (read these)", "h1"),
        ("1. Y5 revenue includes first-cycle mortgage surplus realisation. Underlying recurring revenue is shown separately on every tab.", ""),
        ("2. The 3 cases differ on six drivers: licensee sign-up cadence, mortgages per licensee, Y1 conversion, loan margin, S&P LT return, hedge configuration. Cost ramps adjust accordingly.", ""),
        ("   Conservative ≈ 25% probability mass at or below — Y1 sales miss + slow USA + 6% S&P + 22 bps margin.", ""),
        ("   Likely ≈ 50% probability mass within ±25% — on-plan execution + 8% S&P (defensive vs 10% historical headline) + 25 bps margin.", ""),
        ("   Optimistic ≈ 25% probability mass at or above — strong execution + Pavel-optimised hedge configuration + above-baseline margin.", ""),
        ("   The Likely-case S&P assumption (8%) intentionally below the historical 10% — accounts for fees, sequence risk, and net-of-friction realisable returns over a 30-year horizon. Most analyst frames use 7–8% for forward-looking LT equity.", ""),
        ("3. Y0 represents pre-launch/initial founding; Y1 is the first commercial year.", ""),
        ("4. AI infra cost line is a pitch-realistic estimate, not bound to a specific vendor contract. Update when contracts are signed.", ""),
        ("5. PoC (Probability of Capital shortfall) is NOT in this model — it lives in Futureproof's proprietary financial model (separate Monte Carlo engine). Cite that separately on deck slide 6.", ""),
        ("", ""),
        ("Onboarding fee benchmark — $1.5M per licensee", "h1"),
        ("The $1.5M onboarding fee per Product Issuer (Stream 1 in P&L) was inherited from the prior model with the rationale: 'cost recovery for IT and implementation costs + 30% professional services margin'.", ""),
        ("This is a research-only benchmark check against comparable B2B fintech / regulated-finance SaaS implementation fees. Not company-specific; not a quote.", ""),
        ("", ""),
        ("Comparable fee ranges (industry research)", "h1"),
        ("  Core banking platforms (Mambu, Thought Machine, 10x Banking) — $500K – $5M setup, scale-dependent", ""),
        ("  Lending platforms for institutional rollouts (nCino, Blend, Encompass) — $250K – $2M implementation", ""),
        ("  Insurance core platforms (Guidewire, Duck Creek, Majesco) — $1M – $10M+ full implementations; $1–3M typical mid-tier", ""),
        ("  Wealth / asset-management platforms (Avaloq, Temenos, FNZ) — $500K – $5M+", ""),
        ("  Banking-as-a-Service / new-product launch platforms (Marqeta, Stripe Treasury) — $250K – $1.5M setup", ""),
        ("  API-led narrow integrations (Plaid, Codat type) — $50K – $500K (lower because narrower scope)", ""),
        ("", ""),
        ("Where EPM onboarding sits", "h1"),
        ("EPM onboarding for a Product Issuer involves: (1) tech platform deployment + integration; (2) regulatory mapping for the issuer's jurisdiction; (3) actuarial training on EPM mechanics; (4) issuer-staff training; (5) wholesale-funder linkage; (6) ongoing support transition. Profile is closer to insurance core / wealth-management implementation than to API-led BaaS — placing $1.5M in the middle of the comparable range.", ""),
        ("", ""),
        ("Recommendation", "h1"),
        ("The $1.5M figure is defensible against the comparable set, but should be confirmed against:", ""),
        ("  (a) Actual costed FP onboarding effort — engineering + legal + professional-services hours × loaded rates", ""),
        ("  (b) One real-world reference quote from a current Product Issuer in negotiation", ""),
        ("  (c) CFO sign-off (Wesley Chow) on the cost-recovery + 30% margin framing", ""),
        ("", ""),
        ("Materiality", "h1"),
        ("Onboarding-fee revenue at Y3 (Likely) ≈ $9M (~10% of total revenue). At Y5 ≈ $3M (<1% of total revenue with profit-share realisation). The number could move ±$500K without materially changing the headline P&L — but it is a frequent diligence question, so backing it with a comparable set + cost breakdown is worth the effort.", ""),
        ("", ""),
        ("Pitch deliverables that depend on this model", "h1"),
        ("  docs/capital_raise/deck_outline.md — slides 9, 11, 14 quote numbers from this model.", ""),
        ("  docs/capital_raise/teaser.md — uses Likely-case headline numbers.", ""),
        ("  docs/capital_raise/teaser_vc.md — uses 3-case ranges.", ""),
        ("  docs/capital_raise/faq.md — Q12, Q13, Q14, Q19, Q20 reference these numbers.", ""),
        ("", ""),
    ]

    for i, (text, kind) in enumerate(rows, start=1):
        c = ws.cell(row=i, column=1, value=text)
        if kind == "title":
            c.font = TITLE_FONT
            c.fill = NAVY_FILL
            ws.row_dimensions[i].height = 28
        elif kind == "h1":
            c.font = H2_FONT
            c.fill = LIGHT_FILL
        else:
            c.font = NORMAL_FONT
            c.alignment = Alignment(wrap_text=True, vertical="top")


def build_drivers(wb):
    ws = wb.create_sheet("Drivers")
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 20
    for col in ["C", "D", "E"]:
        ws.column_dimensions[col].width = 16
    ws.column_dimensions["F"].width = 80

    style_title_row(ws, 1, "Drivers — Conservative / Likely / Optimistic", span=6)
    ws.cell(row=2, column=1, value=f"USD/AUD FX: {USD_AUD_FX}").font = NOTE_FONT
    ws.cell(row=3, column=1, value="All driver values in USD unless otherwise noted").font = NOTE_FONT

    headers = ["Driver", "Unit", "Conservative", "Likely", "Optimistic", "Rationale / Notes"]
    for i, h in enumerate(headers):
        c = ws.cell(row=5, column=i + 1, value=h)
        c.font = H2_FONT
        c.fill = LIGHT_FILL

    # Layout: section row, then driver rows for each section
    sections = [
        ("Narrative", [
            ("Case narrative (1 line)", "—",
             CONSERVATIVE.narrative, LIKELY.narrative, OPTIMISTIC.narrative,
             "Each case is internally coherent — drivers move together, not independently."),
        ]),
        ("Sales / Volume", [
            ("Licensees signed UK/AU Y1", "count", CONSERVATIVE.licensees_uk_au[0], LIKELY.licensees_uk_au[0], OPTIMISTIC.licensees_uk_au[0], "Super-fund partnership decisions — the #1 schedule risk."),
            ("Licensees signed UK/AU Y2", "count", CONSERVATIVE.licensees_uk_au[1], LIKELY.licensees_uk_au[1], OPTIMISTIC.licensees_uk_au[1], ""),
            ("Licensees signed UK/AU Y3", "count", CONSERVATIVE.licensees_uk_au[2], LIKELY.licensees_uk_au[2], OPTIMISTIC.licensees_uk_au[2], ""),
            ("Licensees signed UK/AU Y4", "count", CONSERVATIVE.licensees_uk_au[3], LIKELY.licensees_uk_au[3], OPTIMISTIC.licensees_uk_au[3], ""),
            ("Licensees signed UK/AU Y5", "count", CONSERVATIVE.licensees_uk_au[4], LIKELY.licensees_uk_au[4], OPTIMISTIC.licensees_uk_au[4], ""),
            ("Licensees signed USA Y1", "count", CONSERVATIVE.licensees_usa[0], LIKELY.licensees_usa[0], OPTIMISTIC.licensees_usa[0], "USA delayed 1 year vs UK/AU in all cases."),
            ("Licensees signed USA Y2", "count", CONSERVATIVE.licensees_usa[1], LIKELY.licensees_usa[1], OPTIMISTIC.licensees_usa[1], ""),
            ("Licensees signed USA Y3", "count", CONSERVATIVE.licensees_usa[2], LIKELY.licensees_usa[2], OPTIMISTIC.licensees_usa[2], ""),
            ("Licensees signed USA Y4", "count", CONSERVATIVE.licensees_usa[3], LIKELY.licensees_usa[3], OPTIMISTIC.licensees_usa[3], ""),
            ("Licensees signed USA Y5", "count", CONSERVATIVE.licensees_usa[4], LIKELY.licensees_usa[4], OPTIMISTIC.licensees_usa[4], ""),
            ("Mortgages per licensee Y1 target", "count", CONSERVATIVE.mortgages_per_licensee_y1, LIKELY.mortgages_per_licensee_y1, OPTIMISTIC.mortgages_per_licensee_y1, "Penetration into licensee's customer base. Stretchy in Optimistic."),
            ("Mortgages growth p.a.", "%", CONSERVATIVE.mortgages_growth_pa, LIKELY.mortgages_growth_pa, OPTIMISTIC.mortgages_growth_pa, "Year-over-year increase in per-licensee volume."),
            ("Y1 conversion rate (new licensees first year)", "%", CONSERVATIVE.y1_conversion_rate, LIKELY.y1_conversion_rate, OPTIMISTIC.y1_conversion_rate, "First-year ramp for any new licensee."),
        ]),
        ("Pricing (aligned to EpmModelConfig v14a)", [
            ("FP annual margin (FLAT, per v14a)", "bps", CONSERVATIVE.fp_annual_margin_bps, LIKELY.fp_annual_margin_bps, OPTIMISTIC.fp_annual_margin_bps, "v14a baseline = 25 bps. Optimistic uses v14a-OPTIMISED 10 bps (lower drag, higher profit share dollars).") ,
            ("Onboarding fee per licensee", "USD", CONSERVATIVE.onboarding_fee_per_licensee_usd, LIKELY.onboarding_fee_per_licensee_usd, OPTIMISTIC.onboarding_fee_per_licensee_usd, "One-off implementation fee. Same across cases (cost-recovery driven)."),
            ("Capital markets line per licensee p.a.", "USD", CONSERVATIVE.cap_markets_line_per_licensee_usd, LIKELY.cap_markets_line_per_licensee_usd, OPTIMISTIC.cap_markets_line_per_licensee_usd, "New funding line each year per licensee. Optimistic = larger line via strategic anchor."),
            ("Capital markets gross fee", "bps", CONSERVATIVE.cap_markets_fee_bps, LIKELY.cap_markets_fee_bps, OPTIMISTIC.cap_markets_fee_bps, "Third-party arranger fee."),
            ("FP commission on cap-mkts fee", "%", CONSERVATIVE.fp_commission_pct, LIKELY.fp_commission_pct, OPTIMISTIC.fp_commission_pct, "FP share of arranger fee."),
            ("LTV", "%", CONSERVATIVE.ltv, LIKELY.ltv, OPTIMISTIC.ltv, "Regulatory / product cap. Same across cases."),
            ("Avg home value UK/AU", "USD", CONSERVATIVE.home_value_uk_au_usd, LIKELY.home_value_uk_au_usd, OPTIMISTIC.home_value_uk_au_usd, "Targeting higher-value homes (≈AUD 1.8M) in UK/AU."),
            ("Avg home value USA", "USD", CONSERVATIVE.home_value_usa_usd, LIKELY.home_value_usa_usd, OPTIMISTIC.home_value_usa_usd, "USA target market."),
        ]),
        ("Investment & profit share (aligned to v14d Optimised)", [
            ("S&P 500 long-term annual return", "%", CONSERVATIVE.sp500_lt_return, LIKELY.sp500_lt_return, OPTIMISTIC.sp500_lt_return, "Drives K-yr surplus accrual. Conservative uses 6% (below LT avg); Likely 8% (defensive); Optimistic 10% (LT historical)."),
            ("FP profit share at K-yr reset", "%", CONSERVATIVE.profit_share_pct, LIKELY.profit_share_pct, OPTIMISTIC.profit_share_pct, "Per v14d EpmModelConfig: profit_share_pct = 10%. The 50/50 split happens at y30 maturity (beyond 5-yr horizon)."),
            ("Profit-share reset cadence", "years", CONSERVATIVE.profit_share_reset_years, LIKELY.profit_share_reset_years, OPTIMISTIC.profit_share_reset_years, "Per v14d EpmModelConfig: K = 3 years (was 5 years in earlier specs). First reset for Y1-vintage hits at Y4."),
        ]),
        ("Costs", [
            ("FTE Y0", "count", CONSERVATIVE.fte[0], LIKELY.fte[0], OPTIMISTIC.fte[0], "Founding team. Same across cases."),
            ("FTE Y1", "count", CONSERVATIVE.fte[1], LIKELY.fte[1], OPTIMISTIC.fte[1], "Hiring scaled to revenue trajectory."),
            ("FTE Y2", "count", CONSERVATIVE.fte[2], LIKELY.fte[2], OPTIMISTIC.fte[2], ""),
            ("FTE Y3", "count", CONSERVATIVE.fte[3], LIKELY.fte[3], OPTIMISTIC.fte[3], ""),
            ("FTE Y4", "count", CONSERVATIVE.fte[4], LIKELY.fte[4], OPTIMISTIC.fte[4], ""),
            ("FTE Y5", "count", CONSERVATIVE.fte[5], LIKELY.fte[5], OPTIMISTIC.fte[5], "Optimistic adds 17 vs Likely's plateau at 38."),
            ("Avg loaded salary", "USD", CONSERVATIVE.avg_loaded_salary_usd, LIKELY.avg_loaded_salary_usd, OPTIMISTIC.avg_loaded_salary_usd, "Senior-heavy team in regulated finance."),
            ("Salary growth p.a.", "%", CONSERVATIVE.salary_growth_pa, LIKELY.salary_growth_pa, OPTIMISTIC.salary_growth_pa, "Indexation."),
            ("Marketing % of revenue Y1", "%", CONSERVATIVE.marketing_pct_of_rev[1], LIKELY.marketing_pct_of_rev[1], OPTIMISTIC.marketing_pct_of_rev[1], "Conservative spends MORE % to push uphill."),
            ("Marketing % of revenue Y5", "%", CONSERVATIVE.marketing_pct_of_rev[5], LIKELY.marketing_pct_of_rev[5], OPTIMISTIC.marketing_pct_of_rev[5], "Optimistic benefits from organic / inbound."),
            ("Per-FTE overhead p.a.", "USD", CONSERVATIVE.per_fte_overhead_usd, LIKELY.per_fte_overhead_usd, OPTIMISTIC.per_fte_overhead_usd, "Rent, IT, telecom, etc."),
            ("Sales travel per cumulative licensee p.a.", "USD", CONSERVATIVE.sales_travel_per_licensee_usd, LIKELY.sales_travel_per_licensee_usd, OPTIMISTIC.sales_travel_per_licensee_usd, ""),
            ("Legal fees % of recurring rev", "%", CONSERVATIVE.legal_fees_pct_rev, LIKELY.legal_fees_pct_rev, OPTIMISTIC.legal_fees_pct_rev, "Optimistic benefits from scale leverage."),
            ("Actuarial / quant % of rev", "%", CONSERVATIVE.actuarial_pct_rev, LIKELY.actuarial_pct_rev, OPTIMISTIC.actuarial_pct_rev, ""),
            ("Audit % of rev", "%", CONSERVATIVE.audit_pct_rev, LIKELY.audit_pct_rev, OPTIMISTIC.audit_pct_rev, ""),
            ("Platform build Y0 (one-off)", "USD", CONSERVATIVE.platform_build_y0_usd, LIKELY.platform_build_y0_usd, OPTIMISTIC.platform_build_y0_usd, "Same — Y0 build is mostly done; cost is sunk."),
            ("Platform build decay rate", "%", CONSERVATIVE.platform_build_decay, LIKELY.platform_build_decay, OPTIMISTIC.platform_build_decay, "How fast platform-build cost falls Y1+."),
            ("Contingency %", "%", CONSERVATIVE.contingency_pct, LIKELY.contingency_pct, OPTIMISTIC.contingency_pct, "Applied to total opex."),
        ]),
        ("AI infrastructure (NEW vs v10.1)", [
            ("AI infra Y0", "USD", CONSERVATIVE.ai_infra_y0_usd, LIKELY.ai_infra_y0_usd, OPTIMISTIC.ai_infra_y0_usd, "LLM API spend + AI engineering tools. v10.1 had this buried in IT @ $4K/FTE — too low for an AI-first business."),
            ("AI infra growth p.a.", "%", CONSERVATIVE.ai_infra_growth_pa, LIKELY.ai_infra_growth_pa, OPTIMISTIC.ai_infra_growth_pa, "Scales with mortgage volume + customer interactions."),
        ]),
        ("Other", [
            ("Interest rate on cash", "%", CONSERVATIVE.interest_rate_on_cash, LIKELY.interest_rate_on_cash, OPTIMISTIC.interest_rate_on_cash, "T-bill + cash mgmt yield."),
            ("Other income growth p.a.", "x", CONSERVATIVE.other_income_growth, LIKELY.other_income_growth, OPTIMISTIC.other_income_growth, "Captive insurer + ancillary revenue ramp."),
        ]),
    ]

    row = 6
    for section_title, drivers in sections:
        style_section_row(ws, row, section_title, span=6)
        row += 1
        for driver_tuple in drivers:
            label, unit, c_val, l_val, o_val, note = driver_tuple
            ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
            ws.cell(row=row, column=2, value=unit).font = NORMAL_FONT
            for col, val in [(3, c_val), (4, l_val), (5, o_val)]:
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = NORMAL_FONT
                if isinstance(val, float):
                    if "%" in unit or unit in ("x",):
                        cell.number_format = PCT_FMT if "%" in unit else "0.00\"x\""
                    else:
                        cell.number_format = USD_FMT
                elif isinstance(val, int):
                    cell.number_format = CNT_FMT
            note_cell = ws.cell(row=row, column=6, value=note)
            note_cell.font = NOTE_FONT
            note_cell.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
        row += 1

    ws.row_dimensions[6].height = 22  # narrative section header
    # narrative row text wrap
    for col in [3, 4, 5, 6]:
        cell = ws.cell(row=7, column=col)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[7].height = 70


def build_pnl(wb, case: CaseDrivers, pnl: dict):
    sheet_name = f"P&L_{case.name}"
    ws = wb.create_sheet(sheet_name)

    ws.column_dimensions["A"].width = 50
    for col_letter in ["B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col_letter].width = 16

    style_title_row(ws, 1, f"P&L — {case.name} Case (USD)", span=8)
    ws.cell(row=2, column=1, value=case.narrative).font = NOTE_FONT
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    ws.row_dimensions[2].height = 45

    # Year header row
    headers = ["Line item"] + YEARS + ["5-yr Total"]
    for i, h in enumerate(headers):
        c = ws.cell(row=4, column=i + 1, value=h)
        c.font = H2_FONT
        c.fill = LIGHT_FILL
        c.alignment = Alignment(horizontal="right")
    ws.cell(row=4, column=1).alignment = Alignment(horizontal="left")

    row = 5

    def write_line(label, key, fmt=USD_FMT, indent=False, bold=False, fill=None):
        nonlocal row
        c = ws.cell(row=row, column=1, value=label)
        c.font = TOTAL_FONT if bold else NORMAL_FONT
        if indent:
            c.alignment = Alignment(indent=1)
        if fill:
            c.fill = fill
        vals = pnl[key] if key else [None] * 6
        five_yr = sum(v for v in vals[1:] if isinstance(v, (int, float)))
        for i, v in enumerate(vals):
            cell = ws.cell(row=row, column=2 + i, value=v)
            cell.font = TOTAL_FONT if bold else NORMAL_FONT
            cell.number_format = fmt
            if fill:
                cell.fill = fill
        cell = ws.cell(row=row, column=8, value=five_yr)
        cell.font = TOTAL_FONT if bold else NORMAL_FONT
        cell.number_format = fmt
        if fill:
            cell.fill = fill
        row += 1

    def section(text, fill=SECTION_FILL):
        nonlocal row
        style_section_row(ws, row, text, span=8, fill=fill)
        row += 1

    # ---- REVENUE ----
    section("REVENUE")
    write_line("Stream 1 — Onboarding fee (one-off per licensee)", "onboarding", indent=True)
    write_line("Stream 2 — SaaS recurring (loan margin × cumulative book)", "saas", indent=True)
    write_line("Stream 3 — Mortgage surplus (50% of S&P appreciation, 5-yr resets)", "surplus", indent=True)
    write_line("Stream 4 — Capital markets arranger commission", "cap_mkts", indent=True)
    write_line("Stream 5 — Other (interest, captive insurer)", "other", indent=True)
    write_line("Recurring revenue (ex-surplus realisation)", "rev_recurring", bold=True, fill=LIGHT_FILL)
    write_line("Total revenue (incl. surplus realisation)", "rev_total", bold=True, fill=TOTAL_FILL)

    # ---- COSTS ----
    section("OPERATING EXPENSES")
    write_line("Salaries (loaded)", "salaries", indent=True)
    write_line("Per-FTE overheads (rent, IT, telecom, sundry)", "overheads", indent=True)
    write_line("Marketing", "marketing", indent=True)
    write_line("Sales travel", "sales_travel", indent=True)
    write_line("Legal fees", "legal", indent=True)
    write_line("Actuarial / quant / asset consultants", "actuarial", indent=True)
    write_line("Audit / accounting", "audit", indent=True)
    write_line("Platform build & IT development", "platform", indent=True)
    write_line("AI infrastructure (LLM API + tooling)", "ai_infra", indent=True)
    write_line("IP / patents / trademarks", "ip_legal", indent=True)
    write_line("Non-exec director fees", "ned_fees", indent=True)
    write_line("Bank fees & taxes", "bank_fees", indent=True)
    write_line("Total opex (pre-contingency)", "total_opex", bold=True, fill=LIGHT_FILL)
    write_line("Contingency", "contingency", indent=True)
    write_line("Total operating expenses", "total_opex_with_cont", bold=True, fill=TOTAL_FILL)

    # ---- EBITDA ----
    section("EBITDA")
    write_line("EBITDA — recurring (ex-surplus)", "ebitda_recurring", bold=True, fill=EBITDA_FILL)
    write_line("EBITDA — total (incl. surplus)", "ebitda", bold=True, fill=EBITDA_FILL)
    write_line("EBITDA margin — recurring", "ebitda_pct_recurring", fmt=PCT_FMT, indent=True)
    write_line("EBITDA margin — total", "ebitda_pct_total", fmt=PCT_FMT, indent=True)

    # ---- OPERATIONAL METRICS ----
    section("OPERATIONAL METRICS")
    write_line("FTE", "fte", fmt=CNT_FMT, indent=True)
    write_line("Cumulative licensees UK/AU", "cum_lic_uk_au", fmt=CNT_FMT, indent=True)
    write_line("Cumulative licensees USA", "cum_lic_usa", fmt=CNT_FMT, indent=True)
    write_line("New mortgages this year (UK/AU)", "new_mort_uk_au", fmt=CNT_FMT, indent=True)
    write_line("New mortgages this year (USA)", "new_mort_usa", fmt=CNT_FMT, indent=True)
    write_line("Cumulative mortgages on book (UK/AU)", "cum_mort_uk_au", fmt=CNT_FMT, indent=True)
    write_line("Cumulative mortgages on book (USA)", "cum_mort_usa", fmt=CNT_FMT, indent=True)
    write_line("Cumulative mortgages on book (TOTAL)", "cum_mort_total", fmt=CNT_FMT, bold=True, fill=LIGHT_FILL)
    write_line("AUM UK/AU (USD)", "aum_uk_au", fmt=USD_M_FMT, indent=True)
    write_line("AUM USA (USD)", "aum_usa", fmt=USD_M_FMT, indent=True)
    write_line("AUM TOTAL (USD)", "aum_total", fmt=USD_M_FMT, bold=True, fill=LIGHT_FILL)
    write_line("Loan margin (bps)", "loan_margins_bps", fmt=BPS_FMT, indent=True)
    write_line("Opex per AUM (bps)  ◄ AI-leverage punchline", "opex_per_aum_bps", fmt=BPS_FMT, bold=True, fill=EBITDA_FILL)
    write_line("Cumulative recurring cash burn (pre-raise)", "cumulative_cash_burn", indent=True)


def build_summary(wb, results: dict[str, dict]):
    ws = wb.create_sheet("Summary_3Case")
    ws.column_dimensions["A"].width = 38
    for col_letter in "BCDEFGHIJKLMNOPQR":
        ws.column_dimensions[col_letter].width = 12

    style_title_row(ws, 1, "Summary — 3-Case side by side (USD)", span=20)

    # Header structure: Line item | Conservative Y0..Y5 | Likely Y0..Y5 | Optimistic Y0..Y5
    row = 3
    ws.cell(row=row, column=1, value="").fill = LIGHT_FILL
    for i, case in enumerate(CASES):
        start_col = 2 + i * 6
        end_col = start_col + 5
        c = ws.cell(row=row, column=start_col, value=case.name)
        c.font = H2_FONT
        c.fill = LIGHT_FILL
        c.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    row += 1
    ws.cell(row=row, column=1, value="Line item").font = H2_FONT
    ws.cell(row=row, column=1).fill = LIGHT_FILL
    for case_idx in range(3):
        for i, y in enumerate(YEARS):
            c = ws.cell(row=row, column=2 + case_idx * 6 + i, value=y)
            c.font = H2_FONT
            c.fill = LIGHT_FILL
            c.alignment = Alignment(horizontal="right")
    row += 1

    def metric_row(label, key, fmt=USD_FMT, bold=False, fill=None):
        nonlocal row
        c = ws.cell(row=row, column=1, value=label)
        c.font = TOTAL_FONT if bold else NORMAL_FONT
        if fill:
            c.fill = fill
        for case_idx, case in enumerate(CASES):
            vals = results[case.name][key]
            for i, v in enumerate(vals):
                cell = ws.cell(row=row, column=2 + case_idx * 6 + i, value=v)
                cell.font = TOTAL_FONT if bold else NORMAL_FONT
                cell.number_format = fmt
                if fill:
                    cell.fill = fill
        row += 1

    def section(text):
        nonlocal row
        style_section_row(ws, row, text, span=20)
        row += 1

    section("REVENUE")
    metric_row("Recurring revenue", "rev_recurring")
    metric_row("Total revenue (incl. surplus)", "rev_total", bold=True, fill=TOTAL_FILL)

    section("EBITDA")
    metric_row("EBITDA recurring", "ebitda_recurring")
    metric_row("EBITDA total", "ebitda", bold=True, fill=EBITDA_FILL)
    metric_row("EBITDA margin (total)", "ebitda_pct_total", fmt=PCT_FMT)

    section("SCALE")
    metric_row("Cumulative mortgages on book", "cum_mort_total", fmt=CNT_FMT, bold=True)
    metric_row("AUM (USD)", "aum_total", fmt=USD_M_FMT, bold=True)
    metric_row("FTE", "fte", fmt=CNT_FMT)
    metric_row("Opex per AUM (bps)", "opex_per_aum_bps", fmt=BPS_FMT, bold=True, fill=EBITDA_FILL)


def build_au_only(wb, results: dict[str, dict]):
    """AU-only carve-out — UK/AU portion isolated."""
    ws = wb.create_sheet("AU_Only")
    ws.column_dimensions["A"].width = 50
    for col_letter in "BCDEFGHIJKLMNOPQR":
        ws.column_dimensions[col_letter].width = 14

    style_title_row(ws, 1, "AU-Only View — UK/AU portion isolated (USD; AUD in italics)", span=20)
    ws.cell(row=2, column=1, value=(
        "v10.1 collapsed UK/AU into one regional bucket — this tab uses that as a proxy for AU. "
        "When AU/UK split is needed, it requires v10.3 with separate region drivers. "
        "Used on deck slide 7 to show 'AU-only economics' for the AU-led raise narrative."
    )).font = NOTE_FONT
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=20)
    ws.row_dimensions[2].height = 45

    row = 4
    ws.cell(row=row, column=1, value="").fill = LIGHT_FILL
    for i, case in enumerate(CASES):
        start_col = 2 + i * 6
        end_col = start_col + 5
        c = ws.cell(row=row, column=start_col, value=f"{case.name} (UK/AU only)")
        c.font = H2_FONT
        c.fill = LIGHT_FILL
        c.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    row += 1
    for case_idx in range(3):
        for i, y in enumerate(YEARS):
            c = ws.cell(row=row, column=2 + case_idx * 6 + i, value=y)
            c.font = H2_FONT
            c.fill = LIGHT_FILL
            c.alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=1, value="Line item").font = H2_FONT
    ws.cell(row=row, column=1).fill = LIGHT_FILL
    row += 1

    def write_metric(label, key, fmt=USD_FMT, bold=False, fill=None):
        nonlocal row
        c = ws.cell(row=row, column=1, value=label)
        c.font = TOTAL_FONT if bold else NORMAL_FONT
        if fill:
            c.fill = fill
        for case_idx, case in enumerate(CASES):
            vals = results[case.name][key]
            for i, v in enumerate(vals):
                cell = ws.cell(row=row, column=2 + case_idx * 6 + i, value=v)
                cell.font = TOTAL_FONT if bold else NORMAL_FONT
                cell.number_format = fmt
                if fill:
                    cell.fill = fill
        row += 1

    style_section_row(ws, row, "AU/UK volume", span=20); row += 1
    write_metric("Cumulative licensees UK/AU", "cum_lic_uk_au", fmt=CNT_FMT)
    write_metric("New mortgages UK/AU per year", "new_mort_uk_au", fmt=CNT_FMT)
    write_metric("Cumulative mortgages on book UK/AU", "cum_mort_uk_au", fmt=CNT_FMT, bold=True)
    write_metric("AUM UK/AU (USD)", "aum_uk_au", fmt=USD_M_FMT, bold=True, fill=LIGHT_FILL)

    # AUD conversion row for AUM
    style_section_row(ws, row, f"AUD conversion (1 USD = {USD_AUD_FX} AUD)", span=20); row += 1
    c = ws.cell(row=row, column=1, value="AUM UK/AU (AUD)")
    c.font = TOTAL_FONT
    c.fill = LIGHT_FILL
    for case_idx, case in enumerate(CASES):
        vals = results[case.name]["aum_uk_au"]
        for i, v in enumerate(vals):
            cell = ws.cell(row=row, column=2 + case_idx * 6 + i, value=v * USD_AUD_FX)
            cell.font = TOTAL_FONT
            cell.fill = LIGHT_FILL
            cell.number_format = USD_M_FMT
    row += 1


def build_surplus_sensitivity(wb, results: dict[str, dict]):
    ws = wb.create_sheet("Surplus_Sensitivity")
    ws.column_dimensions["A"].width = 55
    for col_letter in "BCDEFG":
        ws.column_dimensions[col_letter].width = 18

    style_title_row(ws, 1, "Surplus Sensitivity — Y5 first-cycle realisation under different S&P returns", span=6)
    ws.cell(row=2, column=1, value=(
        "Independent of case. Shows what Y5 surplus revenue looks like at different S&P long-term return assumptions. "
        "Uses the LIKELY case Y1 mortgage cohort as the base (UK/AU + USA combined). "
        "First sharp investor analyst Q: 'what if 7% instead of 10%?' — point them here."
    )).font = NOTE_FONT
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws.row_dimensions[2].height = 45

    row = 4
    headers = ["S&P 500 LT annual return", "Y1 UK/AU mortgages", "Y1 USA mortgages",
               "Surplus per UK/AU mortgage (USD)", "Surplus per USA mortgage (USD)",
               "Total Y5 surplus realisation (USD)"]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=i + 1, value=h)
        c.font = H2_FONT
        c.fill = LIGHT_FILL
    row += 1

    likely_y1_uk_au = results["Likely"]["new_mort_uk_au"][1]
    likely_y1_usa = results["Likely"]["new_mort_usa"][1]
    loan_size_uk_au = LIKELY.ltv * LIKELY.home_value_uk_au_usd
    loan_size_usa = LIKELY.ltv * LIKELY.home_value_usa_usd

    K = LIKELY.profit_share_reset_years
    for sp_return in [0.05, 0.07, 0.10, 0.12]:
        s_uk = loan_size_uk_au * ((1 + sp_return) ** K - 1) * LIKELY.profit_share_pct
        s_us = loan_size_usa * ((1 + sp_return) ** K - 1) * LIKELY.profit_share_pct
        total = likely_y1_uk_au * s_uk + likely_y1_usa * s_us

        ws.cell(row=row, column=1, value=f"{sp_return:.0%}")
        ws.cell(row=row, column=2, value=likely_y1_uk_au).number_format = CNT_FMT
        ws.cell(row=row, column=3, value=likely_y1_usa).number_format = CNT_FMT
        ws.cell(row=row, column=4, value=s_uk).number_format = USD_FMT
        ws.cell(row=row, column=5, value=s_us).number_format = USD_FMT
        cell = ws.cell(row=row, column=6, value=total)
        cell.number_format = USD_M_FMT
        cell.font = TOTAL_FONT
        if abs(sp_return - 0.10) < 1e-6:
            cell.fill = TOTAL_FILL
        row += 1

    ws.cell(row=row + 1, column=1, value=(
        "Read: at 7% S&P (vs 10% Likely), Y5 surplus is ~40% lower. The recurring-revenue economics "
        "(SaaS + onboarding + cap-markets) are unaffected; only the surplus stream moves."
    )).font = NOTE_FONT
    ws.cell(row=row + 1, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=6)


def build_runway(wb, results: dict[str, dict]):
    ws = wb.create_sheet("Runway_vs_Raise")
    ws.column_dimensions["A"].width = 50
    for col_letter in "BCDEFGHIJKLM":
        ws.column_dimensions[col_letter].width = 16

    style_title_row(ws, 1, "Runway vs Raise — does the raise size cover the burn?", span=12)
    ws.cell(row=2, column=1, value=(
        f"FX assumption: 1 USD = {USD_AUD_FX} AUD. Cumulative cash burn = sum of recurring-EBITDA Y0+Y1+...+YN "
        "(NOT including surplus, which doesn't realise until Y5). "
        "Compare against raise scenarios A$5M / A$7.5M / A$10M, converted to USD."
    )).font = NOTE_FONT
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)
    ws.row_dimensions[2].height = 45

    row = 4
    headers = ["Case", "Cum burn Y0-Y2 (USD)", "Cum burn Y0-Y2 (AUD)",
               "Y2 EBITDA (rec, USD)", "Y3 EBITDA (rec, USD)",
               "A$5M raise: covers?", "A$7.5M raise: covers?", "A$10M raise: covers?",
               "Recommendation"]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=i + 1, value=h)
        c.font = H2_FONT
        c.fill = LIGHT_FILL
        c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 40
    row += 1

    raise_scenarios_aud = [5_000_000, 7_500_000, 10_000_000]
    raise_scenarios_usd = [r / USD_AUD_FX for r in raise_scenarios_aud]

    for case in CASES:
        burn_y0 = -results[case.name]["ebitda_recurring"][0]
        burn_y1 = -results[case.name]["ebitda_recurring"][1]
        burn_y2 = -results[case.name]["ebitda_recurring"][2]
        cum_burn_y0_y2 = max(0, burn_y0) + max(0, burn_y1) + max(0, burn_y2)
        ebitda_y2 = results[case.name]["ebitda_recurring"][2]
        ebitda_y3 = results[case.name]["ebitda_recurring"][3]

        verdicts = []
        for raise_usd in raise_scenarios_usd:
            buffer = raise_usd - cum_burn_y0_y2
            buffer_pct = buffer / raise_usd if raise_usd > 0 else 0
            if buffer < 0:
                verdicts.append(f"NO — short {-buffer/1e6:.1f}M USD")
            elif buffer_pct < 0.20:
                verdicts.append(f"TIGHT ({buffer_pct:.0%} buffer)")
            else:
                verdicts.append(f"OK ({buffer_pct:.0%} buffer)")

        rec = "A$10M" if cum_burn_y0_y2 / USD_AUD_FX * 1.25 > 7_500_000 else "A$7.5M target / A$5M minimum"

        vals = [case.name, cum_burn_y0_y2, cum_burn_y0_y2 * USD_AUD_FX,
                ebitda_y2, ebitda_y3,
                verdicts[0], verdicts[1], verdicts[2], rec]
        for i, v in enumerate(vals):
            cell = ws.cell(row=row, column=i + 1, value=v)
            cell.font = NORMAL_FONT
            if isinstance(v, (int, float)):
                cell.number_format = USD_FMT
            if i == 0:
                cell.font = TOTAL_FONT
                cell.fill = LIGHT_FILL
        row += 1

    row += 2
    style_section_row(ws, row, "Raise scenarios — AUD to USD", span=12)
    row += 1
    headers2 = ["Raise (AUD)", "Raise (USD)"]
    for i, h in enumerate(headers2):
        c = ws.cell(row=row, column=i + 1, value=h)
        c.font = H2_FONT
        c.fill = LIGHT_FILL
    row += 1
    for r_aud, r_usd in zip(raise_scenarios_aud, raise_scenarios_usd):
        ws.cell(row=row, column=1, value=r_aud).number_format = USD_FMT
        ws.cell(row=row, column=2, value=r_usd).number_format = USD_FMT
        row += 1


def build_ai_cost_detail(wb, results: dict[str, dict]):
    ws = wb.create_sheet("AI_Cost_Detail")
    ws.column_dimensions["A"].width = 50
    for col_letter in "BCDEFGHIJKLMNOP":
        ws.column_dimensions[col_letter].width = 14

    style_title_row(ws, 1, "AI Cost Detail — operating-leverage thesis quantified", span=16)
    ws.cell(row=2, column=1, value=(
        "Slide 9 of the deck claims AI is operating leverage. This tab shows the actual AI infra spend per case per year, "
        "as % of total opex, and the implied 'AI-absorbed FTE-equivalents' relative to a comparable specialty insurer "
        "(benchmark: 200 FTE per A$1B AUM, vs FutureProof's plan of ~38)."
    )).font = NOTE_FONT
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=16)
    ws.row_dimensions[2].height = 60

    row = 4
    ws.cell(row=row, column=1, value="").fill = LIGHT_FILL
    for i, case in enumerate(CASES):
        start_col = 2 + i * 6
        end_col = start_col + 5
        c = ws.cell(row=row, column=start_col, value=case.name)
        c.font = H2_FONT
        c.fill = LIGHT_FILL
        c.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    row += 1
    ws.cell(row=row, column=1, value="Line item").font = H2_FONT
    ws.cell(row=row, column=1).fill = LIGHT_FILL
    for case_idx in range(3):
        for i, y in enumerate(YEARS):
            c = ws.cell(row=row, column=2 + case_idx * 6 + i, value=y)
            c.font = H2_FONT
            c.fill = LIGHT_FILL
            c.alignment = Alignment(horizontal="right")
    row += 1

    def write_row(label, key, fmt=USD_FMT, transform=None, bold=False, fill=None):
        nonlocal row
        c = ws.cell(row=row, column=1, value=label)
        c.font = TOTAL_FONT if bold else NORMAL_FONT
        if fill:
            c.fill = fill
        for case_idx, case in enumerate(CASES):
            vals = results[case.name][key]
            if transform:
                vals = transform(vals, results[case.name])
            for i, v in enumerate(vals):
                cell = ws.cell(row=row, column=2 + case_idx * 6 + i, value=v)
                cell.font = TOTAL_FONT if bold else NORMAL_FONT
                cell.number_format = fmt
                if fill:
                    cell.fill = fill
        row += 1

    write_row("AI infrastructure cost (USD)", "ai_infra", bold=True, fill=LIGHT_FILL)
    write_row("AI as % of total opex", "ai_infra", fmt=PCT_FMT,
              transform=lambda ai, c: [a/o if o > 0 else 0 for a, o in zip(ai, c["total_opex_with_cont"])])
    write_row("FTE on plan", "fte", fmt=CNT_FMT)
    write_row("AUM TOTAL (USD)", "aum_total", fmt=USD_M_FMT)

    # Comp benchmark row
    style_section_row(ws, row, "Benchmark — Comparable specialty insurer / annuity provider", span=16)
    row += 1
    write_row("FutureProof FTE per $1B AUM", "fte", fmt=CNT_FMT,
              transform=lambda fte, c: [f / max(a/1e9, 0.001) if a > 0 else None for f, a in zip(fte, c["aum_total"])])

    # Static benchmark (200 FTE per $1B AUM)
    c = ws.cell(row=row, column=1, value="Specialty insurer benchmark (FTE per $1B AUM)")
    c.font = NORMAL_FONT
    for case_idx in range(3):
        for i in range(6):
            cell = ws.cell(row=row, column=2 + case_idx * 6 + i, value=200)
            cell.number_format = CNT_FMT
    row += 1

    c = ws.cell(row=row, column=1, value="◄ AI-absorbed FTE-equivalents")
    c.font = TOTAL_FONT
    c.fill = EBITDA_FILL
    for case_idx, case in enumerate(CASES):
        aums = results[case.name]["aum_total"]
        ftes = results[case.name]["fte"]
        for i in range(6):
            if aums[i] > 0:
                expected_fte = 200 * (aums[i] / 1e9)
                absorbed = max(0, expected_fte - ftes[i])
            else:
                absorbed = 0
            cell = ws.cell(row=row, column=2 + case_idx * 6 + i, value=absorbed)
            cell.font = TOTAL_FONT
            cell.fill = EBITDA_FILL
            cell.number_format = CNT_FMT
    row += 1

    ws.cell(row=row + 1, column=1, value=(
        "Read: by Y3 in the Likely case, the AI-ops layer is absorbing ~250+ FTE-equivalents of work that a "
        "comparable specialty insurer would carry. This is the operating-leverage thesis on slide 9 — "
        "quantified, with the benchmark cited."
    )).font = NOTE_FONT
    ws.cell(row=row + 1, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=16)


# ===========================================================================
# VC-style Valuation: Y6-Y10 extrapolation, DCF, VC Method, Comp multiples
# ===========================================================================

def compute_extrapolation(case: CaseDrivers, pnl: dict) -> dict:
    """Project Y6-Y10 from Y5 base for any case.

    Approach:
    - Recurring revenue grows at declining rate as the platform matures.
    - Profit-share is realised by vintage every 5 yrs. Y6 = Y2 vintage; Y7 = Y3 vintage; etc.
      Y10 = Y5 vintage + Y1 second-reset.
    - EBITDA margin held at scale-appropriate level. Conservative slightly lower because the
      Y0-Y5 ramp left less recurring scale to absorb Y6+ opex.
    """
    out = {}

    # Growth rates by case (Conservative slower; Optimistic faster)
    growth_by_case = {
        "Conservative": [0.22, 0.20, 0.18, 0.16, 0.14],
        "Likely":       [0.28, 0.24, 0.20, 0.17, 0.15],
        "Optimistic":   [0.40, 0.32, 0.26, 0.22, 0.18],
    }
    margin_by_case = {
        "Conservative": [0.78, 0.80, 0.82, 0.83, 0.84],
        "Likely":       [0.85, 0.87, 0.88, 0.89, 0.90],
        "Optimistic":   [0.90, 0.91, 0.92, 0.92, 0.93],
    }
    rec_growth = growth_by_case[case.name]
    margin_y6_y10 = margin_by_case[case.name]

    # --- Recurring revenue extrapolation ---
    rec_y5 = pnl["rev_recurring"][5]
    rec = [rec_y5]
    for g in rec_growth:
        rec.append(rec[-1] * (1 + g))
    rec_y6_y10 = rec[1:]

    # --- Profit-share by year (3-yr reset cadence per v14d) ---
    K = case.profit_share_reset_years
    surplus_per_uk_au = case.home_value_uk_au_usd * case.ltv * ((1 + case.sp500_lt_return) ** K - 1) * case.profit_share_pct
    surplus_per_usa = case.home_value_usa_usd * case.ltv * ((1 + case.sp500_lt_return) ** K - 1) * case.profit_share_pct

    new_uk = pnl["new_mort_uk_au"]
    new_us = pnl["new_mort_usa"]

    # Extrapolate Y6 / Y7 vintages (used for Y9 / Y10 first-reset realisations)
    new_uk_extrap = list(new_uk) + [new_uk[5] * 1.08, new_uk[5] * 1.16]  # Y6, Y7
    new_us_extrap = list(new_us) + [new_us[5] * 1.08, new_us[5] * 1.16]

    # Within Y6-Y10 window, profit share comes from any vintage hitting a K-yr reset.
    # Vintage v hits resets at v+K, v+2K, v+3K, ...
    profit_share_y6_y10 = [0, 0, 0, 0, 0]  # Y6..Y10
    for vintage_y in range(1, 8):  # Y1..Y7 vintages may contribute
        reset_n = 1
        while True:
            reset_year = vintage_y + reset_n * K
            if reset_year > 10:
                break
            if 6 <= reset_year <= 10:
                ps = (
                    new_uk_extrap[vintage_y] * surplus_per_uk_au
                    + new_us_extrap[vintage_y] * surplus_per_usa
                )
                profit_share_y6_y10[reset_year - 6] += ps
            reset_n += 1

    # --- Total revenue ---
    rev_total_y6_y10 = [rec + ps for rec, ps in zip(rec_y6_y10, profit_share_y6_y10)]

    # --- EBITDA ---
    ebitda_y6_y10 = [r * m for r, m in zip(rev_total_y6_y10, margin_y6_y10)]

    # --- Normalised EBITDA (smooth profit-share by 0.6 factor for valuation multiples) ---
    normalised_ebitda_y6_y10 = [
        rec * m * 0.95 + (ps * 0.6) * m
        for rec, m, ps in zip(rec_y6_y10, margin_y6_y10, profit_share_y6_y10)
    ]

    # --- FTE & AUM extrapolation ---
    fte_y5 = case.fte[5]
    fte_y6_y10 = [fte_y5 + i for i in range(1, 6)]
    aum_y5 = pnl["aum_total"][5]
    aum_growth_by_case = {
        "Conservative": [0.20, 0.18, 0.16, 0.15, 0.14],
        "Likely":       [0.25, 0.22, 0.20, 0.18, 0.16],
        "Optimistic":   [0.35, 0.30, 0.26, 0.22, 0.20],
    }
    aum_growth = aum_growth_by_case[case.name]
    aum = [aum_y5]
    for g in aum_growth:
        aum.append(aum[-1] * (1 + g))
    aum_y6_y10 = aum[1:]

    out["years"] = ["Year 6", "Year 7", "Year 8", "Year 9", "Year 10"]
    out["recurring"] = rec_y6_y10
    out["profit_share"] = profit_share_y6_y10
    out["rev_total"] = rev_total_y6_y10
    out["ebitda"] = ebitda_y6_y10
    out["normalised_ebitda"] = normalised_ebitda_y6_y10
    out["margin"] = margin_y6_y10
    out["fte"] = fte_y6_y10
    out["aum"] = aum_y6_y10
    return out


def compute_dcf(pnl: dict, extrap: dict, wacc: float, terminal_growth: float, tax_rate: float) -> dict:
    """10-year DCF with terminal value (Gordon growth) using NORMALISED EBITDA.

    Free cash flow = Normalised EBITDA × (1 - tax). Asset-light: capex and working-capital changes assumed minimal.
    Discounting is mid-year convention (typical in VC/banking analyst models).
    """
    # Build full 10-year normalised EBITDA series (Y1-Y10)
    ebitda_y1_y5 = list(pnl["ebitda_recurring"][1:6])  # use recurring for Y1-Y5 (less lumpy)
    # For Y5 specifically, add a smoothed slice of the profit share
    ebitda_y5_smoothed = pnl["ebitda_recurring"][5] + pnl["surplus"][5] * 0.6 * 0.91
    ebitda_y1_y5[4] = ebitda_y5_smoothed
    ebitda_y6_y10 = extrap["normalised_ebitda"]
    ebitda_full = ebitda_y1_y5 + ebitda_y6_y10  # Y1..Y10

    fcf_full = [e * (1 - tax_rate) for e in ebitda_full]

    # PV of explicit period (mid-year discounting)
    pv_fcf = []
    for y, f in enumerate(fcf_full, start=1):
        pv = f / ((1 + wacc) ** (y - 0.5))
        pv_fcf.append(pv)

    npv_explicit = sum(pv_fcf)

    # Terminal value (Gordon growth on Y10 FCF)
    fcf_y11 = fcf_full[-1] * (1 + terminal_growth)
    tv = fcf_y11 / (wacc - terminal_growth) if wacc > terminal_growth else 0
    pv_tv = tv / ((1 + wacc) ** 9.5)  # discount from Y10 mid-year

    enterprise_value = npv_explicit + pv_tv

    return {
        "ebitda_full": ebitda_full,
        "fcf_full": fcf_full,
        "pv_fcf": pv_fcf,
        "npv_explicit": npv_explicit,
        "terminal_value_undiscounted": tv,
        "pv_terminal_value": pv_tv,
        "enterprise_value": enterprise_value,
        "wacc": wacc,
        "terminal_growth": terminal_growth,
        "tax_rate": tax_rate,
    }


def compute_vc_method(pnl: dict, extrap: dict) -> dict:
    """VC Method: project exit value, discount at required IRR, compute implied today's valuation."""
    scenarios = []

    # Scenario A: Y7 strategic acquisition by life insurer / super-fund consolidator
    rev_y7 = extrap["rev_total"][1]  # Y7
    ebitda_y7_norm = extrap["normalised_ebitda"][1]
    exit_a_low = rev_y7 * 4   # 4x revenue (specialty insurance multiple)
    exit_a_high = ebitda_y7_norm * 12  # 12x EBITDA
    scenarios.append({
        "name": "Y7 Strategic Acquisition",
        "year": 7,
        "rev_at_exit": rev_y7,
        "ebitda_at_exit_norm": ebitda_y7_norm,
        "exit_value_low": exit_a_low,
        "exit_value_high": exit_a_high,
    })

    # Scenario B: Y10 IPO at scale
    rev_y10 = extrap["rev_total"][4]
    ebitda_y10_norm = extrap["normalised_ebitda"][4]
    exit_b_low = rev_y10 * 5
    exit_b_high = ebitda_y10_norm * 15
    scenarios.append({
        "name": "Y10 IPO at Scale",
        "year": 10,
        "rev_at_exit": rev_y10,
        "ebitda_at_exit_norm": ebitda_y10_norm,
        "exit_value_low": exit_b_low,
        "exit_value_high": exit_b_high,
    })

    # For each scenario, work back to today's implied valuation at three required IRR levels
    irr_targets = [0.30, 0.40, 0.50]
    for s in scenarios:
        s["implied_today_low"] = {irr: s["exit_value_low"] / ((1 + irr) ** s["year"]) for irr in irr_targets}
        s["implied_today_high"] = {irr: s["exit_value_high"] / ((1 + irr) ** s["year"]) for irr in irr_targets}

    return {"scenarios": scenarios, "irr_targets": irr_targets}


def compute_metrics(pnl: dict, extrap: dict) -> dict:
    """VC dashboard metrics."""
    metrics = {}
    fte = pnl["fte"]
    rev_rec = pnl["rev_recurring"]
    rev_total = pnl["rev_total"]
    ebitda_rec = pnl["ebitda_recurring"]
    ebitda_total = pnl["ebitda"]
    aum = pnl["aum_total"]
    cum_lic = [u + s for u, s in zip(pnl["cum_lic_uk_au"], pnl["cum_lic_usa"])]

    # Revenue per FTE
    metrics["rev_per_fte"] = [r / f if f > 0 else 0 for r, f in zip(rev_rec, fte)]

    # AUM per FTE
    metrics["aum_per_fte"] = [a / f if f > 0 else 0 for a, f in zip(aum, fte)]

    # Revenue per cumulative licensee (B2B customer)
    metrics["rev_per_licensee"] = [r / l if l > 0 else 0 for r, l in zip(rev_total, cum_lic)]

    # Rule of 40 (revenue growth + EBITDA margin)
    rev_growth = [0]
    for y in range(1, 6):
        if rev_rec[y - 1] > 0:
            rev_growth.append((rev_rec[y] - rev_rec[y - 1]) / rev_rec[y - 1])
        else:
            rev_growth.append(None)
    metrics["rev_growth"] = rev_growth
    metrics["ebitda_margin_recurring"] = pnl["ebitda_pct_recurring"]
    metrics["rule_of_40"] = [
        (g + m) if g is not None else None
        for g, m in zip(rev_growth, pnl["ebitda_pct_recurring"])
    ]

    # LTV per B2B customer (5-yr undiscounted) — total revenue / Y5 cumulative licensees
    y5_total_rev = sum(rev_total[1:6])
    y5_cum_lic = cum_lic[5]
    metrics["ltv_per_customer_5yr"] = y5_total_rev / y5_cum_lic if y5_cum_lic > 0 else 0
    metrics["cac_per_customer"] = 1_500_000  # from Drivers / briefing-deck assumption
    metrics["ltv_cac_ratio"] = metrics["ltv_per_customer_5yr"] / metrics["cac_per_customer"]

    # Cash payback (years to recoup CAC from a single customer)
    rev_per_cust_yr = metrics["rev_per_licensee"][3]  # at Y3 steady state
    metrics["cash_payback_years"] = metrics["cac_per_customer"] / rev_per_cust_yr if rev_per_cust_yr > 0 else 0

    return metrics


# ---------------------------------------------------------------------------
# Tab builders for new tabs
# ---------------------------------------------------------------------------

def build_extrapolation(wb, results: dict, extrap_by_case: dict):
    """Y6-Y10 extrapolation, 3 cases stacked vertically."""
    ws = wb.create_sheet("Y6_to_Y10")
    ws.column_dimensions["A"].width = 48
    for c in "BCDEFGHIJK":
        ws.column_dimensions[c].width = 13

    style_title_row(ws, 1, "Y6–Y10 Extrapolation — All 3 Cases (USD)", span=11)
    ws.cell(row=2, column=1, value=(
        "Extension of the 5-year cases to Year 10 for valuation. Each case has its own growth + margin assumptions. "
        "Profit-share is vintage-driven: Y(N) realisation = mortgages from the Y(N-5) vintage. "
        "NORMALISED EBITDA smooths profit-share lumpiness for valuation multiples (multiply current-year profit-share by 0.6 smoothing factor)."
    )).font = NOTE_FONT
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)
    ws.row_dimensions[2].height = 50

    row = 4

    for case in CASES:
        pnl = results[case.name]
        extrap = extrap_by_case[case.name]

        # Case header
        case_color = {"Conservative": "F4CCCC", "Likely": "D9E1F2", "Optimistic": "C6E0B4"}[case.name]
        c = ws.cell(row=row, column=1, value=f"  {case.name.upper()} CASE")
        c.font = H1_FONT
        c.fill = SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
        ws.row_dimensions[row].height = 22
        row += 1

        # Year header
        headers = ["Line item", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"] + extrap["years"]
        for i, h in enumerate(headers):
            c = ws.cell(row=row, column=i + 1, value=h)
            c.font = H2_FONT
            c.fill = LIGHT_FILL
            c.alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
        row += 1

        def write_row(label, y1_y5_vals, y6_y10_vals, fmt=USD_M_FMT, bold=False, fill=None):
            nonlocal row
            c = ws.cell(row=row, column=1, value=label)
            c.font = TOTAL_FONT if bold else NORMAL_FONT
            if fill:
                c.fill = fill
            all_vals = list(y1_y5_vals) + list(y6_y10_vals)
            for i, v in enumerate(all_vals):
                cell = ws.cell(row=row, column=2 + i, value=v)
                cell.font = TOTAL_FONT if bold else NORMAL_FONT
                cell.number_format = fmt
                if fill:
                    cell.fill = fill
            row += 1

        write_row("Recurring revenue", pnl["rev_recurring"][1:6], extrap["recurring"])
        write_row("Profit-share realisation", [0, 0, 0, 0, pnl["surplus"][5]], extrap["profit_share"])
        write_row("Total revenue", pnl["rev_total"][1:6], extrap["rev_total"], bold=True, fill=TOTAL_FILL)
        write_row("EBITDA — total", pnl["ebitda"][1:6], extrap["ebitda"], bold=True, fill=EBITDA_FILL)
        write_row("EBITDA — NORMALISED", pnl["ebitda_recurring"][1:6], extrap["normalised_ebitda"], fill=LIGHT_FILL)
        write_row("AUM (USD)", pnl["aum_total"][1:6], extrap["aum"], fmt=USD_M_FMT)
        row += 1  # blank line between cases

    ws.cell(row=row + 1, column=1, value=(
        "NOTE on profit-share lumpiness: Profit-share realises every 5 years per vintage. Y6 = Y2 vintage; Y10 = Y6 vintage + Y1 second-reset. "
        "The NORMALISED EBITDA line damps the lumpy profit-share component (×0.6) — use this line for valuation multiples, not raw EBITDA."
    )).font = NOTE_FONT
    ws.cell(row=row + 1, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=11)


def build_valuation(wb, results: dict, extrap_by_case: dict, dcf_by_case: dict, vc_by_case: dict):
    """3-case DCF + VC Method + Comparable multiples."""
    ws = wb.create_sheet("Valuation_VC")
    ws.column_dimensions["A"].width = 48
    for c in "BCDEFGHIJ":
        ws.column_dimensions[c].width = 16

    style_title_row(ws, 1, "VC-Style Valuation — 3 Cases (DCF + VC Method + Comparable Multiples)", span=10)
    ws.cell(row=2, column=1, value=(
        "Three-lens valuation pack across all three cases. "
        "(1) DCF — 10-year explicit + Gordon terminal, computed for each case. "
        "(2) VC Method — exit value / required IRR → today's implied valuation. "
        "(3) Comparable multiples (research-only ranges; update with primary-source citations before diligence)."
    )).font = NOTE_FONT
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
    ws.row_dimensions[2].height = 45

    row = 4

    # ----- DCF SUMMARY (3 cases side-by-side) -----
    style_section_row(ws, row, "1a. DCF Summary — 3 cases (WACC 14%, terminal g 3%, tax 30%)", span=10); row += 1

    headers = ["Line item", "Conservative", "Likely", "Optimistic", "Notes"]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=i + 1, value=h)
        c.font = H2_FONT
        c.fill = LIGHT_FILL
    ws.column_dimensions["E"].width = 50
    row += 1

    def dcf_row(label, key_fn, fmt=USD_M_FMT, note="", bold=False, fill=None):
        nonlocal row
        c = ws.cell(row=row, column=1, value=label)
        c.font = TOTAL_FONT if bold else NORMAL_FONT
        if fill:
            c.fill = fill
        for i, case in enumerate(CASES):
            v = key_fn(dcf_by_case[case.name], extrap_by_case[case.name])
            cell = ws.cell(row=row, column=2 + i, value=v)
            cell.font = TOTAL_FONT if bold else NORMAL_FONT
            cell.number_format = fmt
            if fill:
                cell.fill = fill
        c2 = ws.cell(row=row, column=5, value=note)
        c2.font = NOTE_FONT
        c2.alignment = Alignment(wrap_text=True)
        row += 1

    dcf_row("Y10 Total revenue", lambda d, e: e["rev_total"][-1], note="Vintage-driven profit-share + recurring growth.")
    dcf_row("Y10 Normalised EBITDA", lambda d, e: e["normalised_ebitda"][-1], note="Profit-share smoothed ×0.6 for steady-state framing.")
    dcf_row("NPV of explicit 10-yr period", lambda d, e: d["npv_explicit"], fill=LIGHT_FILL)
    dcf_row("Terminal value (undiscounted)", lambda d, e: d["terminal_value_undiscounted"], note="Gordon growth at 3%; could substitute exit multiple.")
    dcf_row("Present value of terminal", lambda d, e: d["pv_terminal_value"], fill=LIGHT_FILL)
    dcf_row("ENTERPRISE VALUE (DCF)", lambda d, e: d["enterprise_value"], bold=True, fill=TOTAL_FILL, note="Discounted FCF + PV of terminal.")
    row += 2

    # ----- DCF SENSITIVITY (Likely only — central case) -----
    style_section_row(ws, row, "1b. DCF Sensitivity — Likely Case · WACC × Terminal Growth (Enterprise Value, USD M)", span=10); row += 1

    likely_pnl = results["Likely"]
    likely_extrap = extrap_by_case["Likely"]
    waccs = [0.10, 0.12, 0.14, 0.16, 0.18]
    tgrowths = [0.02, 0.025, 0.03, 0.035, 0.04]

    ws.cell(row=row, column=1, value="WACC \\ Terminal g").font = H2_FONT
    ws.cell(row=row, column=1).fill = LIGHT_FILL
    for i, g in enumerate(tgrowths):
        c = ws.cell(row=row, column=2 + i, value=g)
        c.number_format = PCT_FMT
        c.font = H2_FONT
        c.fill = LIGHT_FILL
    row += 1

    for w in waccs:
        ws.cell(row=row, column=1, value=w).number_format = PCT_FMT
        ws.cell(row=row, column=1).font = H2_FONT
        ws.cell(row=row, column=1).fill = LIGHT_FILL
        for i, g in enumerate(tgrowths):
            d = compute_dcf(likely_pnl, likely_extrap, w, g, 0.30)
            cell = ws.cell(row=row, column=2 + i, value=d["enterprise_value"])
            cell.number_format = USD_M_FMT
            if abs(w - 0.14) < 1e-6 and abs(g - 0.03) < 1e-6:
                cell.fill = TOTAL_FILL
                cell.font = TOTAL_FONT
        row += 1

    row += 2

    # ----- VC METHOD (3 cases) -----
    style_section_row(ws, row, "2. VC Method — exit value / required IRR → today's implied valuation (3 cases)", span=10); row += 1
    ws.cell(row=row, column=1, value=(
        "For each case: Y7 strategic-acquisition exit (revenue × 4 OR EBITDA × 12) and Y10 IPO exit (revenue × 5 OR EBITDA × 15), "
        "discounted at three required IRR levels (30 / 40 / 50%) to arrive at today's implied valuation."
    )).font = NOTE_FONT
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 2

    for case in CASES:
        vc_method = vc_by_case[case.name]
        irrs = vc_method["irr_targets"]

        # Case header
        c = ws.cell(row=row, column=1, value=f"  {case.name.upper()} CASE")
        c.font = H1_FONT
        c.fill = SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        ws.row_dimensions[row].height = 20
        row += 1

        for s in vc_method["scenarios"]:
            ws.cell(row=row, column=1, value=s["name"]).font = TOTAL_FONT
            ws.cell(row=row, column=1).fill = LIGHT_FILL
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
            row += 1

            ws.cell(row=row, column=1, value=f"Revenue at Y{s['year']}").font = NORMAL_FONT
            cell = ws.cell(row=row, column=2, value=s["rev_at_exit"]); cell.number_format = USD_M_FMT
            ws.cell(row=row, column=3, value="Normalised EBITDA").font = NORMAL_FONT
            cell = ws.cell(row=row, column=4, value=s["ebitda_at_exit_norm"]); cell.number_format = USD_M_FMT
            ws.cell(row=row, column=5, value="Exit (low — rev mult)").font = NORMAL_FONT
            cell = ws.cell(row=row, column=6, value=s["exit_value_low"]); cell.number_format = USD_M_FMT
            ws.cell(row=row, column=7, value="Exit (high — EBITDA mult)").font = NORMAL_FONT
            cell = ws.cell(row=row, column=8, value=s["exit_value_high"]); cell.number_format = USD_M_FMT
            row += 1

            # IRR grid
            ws.cell(row=row, column=1, value="Required IRR →").font = NORMAL_FONT
            for i, irr in enumerate(irrs):
                c = ws.cell(row=row, column=2 + i, value=irr)
                c.number_format = PCT_FMT
                c.font = H2_FONT
            row += 1

            ws.cell(row=row, column=1, value="Today's NPV (low exit)").font = NORMAL_FONT
            for i, irr in enumerate(irrs):
                cell = ws.cell(row=row, column=2 + i, value=s["implied_today_low"][irr])
                cell.number_format = USD_M_FMT
            row += 1

            ws.cell(row=row, column=1, value="Today's NPV (high exit)").font = TOTAL_FONT
            for i, irr in enumerate(irrs):
                cell = ws.cell(row=row, column=2 + i, value=s["implied_today_high"][irr])
                cell.number_format = USD_M_FMT
                cell.font = TOTAL_FONT
                cell.fill = LIGHT_FILL
            row += 1

        row += 1

    ws.cell(row=row, column=1, value=(
        "Read: at $5M raise size, the implied today's-NPV substantially exceeds the round size in every case at every IRR — meaning the round is comfortably underwriteable on VC-method math even at the most demanding 50% IRR target on Conservative-case assumptions."
    )).font = NOTE_FONT
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 3

    # ----- COMPARABLE MULTIPLES (case-independent) -----
    style_section_row(ws, row, "3. Comparable Multiples — research-only ranges; update before diligence", span=10); row += 1

    ws.cell(row=row, column=1, value=(
        "Three category lenses are relevant for FP. RESEARCH-ONLY ranges sourced from public-knowledge analyst frames. "
        "Update with primary-source citations (Forrester, IDC, Celent, Pitchbook) before any investor diligence call."
    )).font = NOTE_FONT
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 2

    headers = ["Category", "Comp companies", "EV / Revenue", "EV / EBITDA", "EV / AUM (bps)", "Notes"]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=i + 1, value=h)
        c.font = H2_FONT
        c.fill = LIGHT_FILL
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["F"].width = 36
    row += 1

    comps = [
        ("Specialty insurance / annuity", "Challenger, Resolution Life, Generation Life, Athene", "1.5–3×", "8–14×", "10–25 bps", "Closest economic comparator. Stable cashflows, regulated, AUM-linked revenue."),
        ("Specialty / non-bank lending", "Heartland Group, Resimac, PepperMoney, Finance of America", "1–2× book", "6–10×", "20–60 bps", "Lower multiple — credit risk concerns. We're explicitly NOT this; useful framing to differentiate."),
        ("Insurance / fintech SaaS platforms", "Guidewire, Duck Creek, FNZ, nCino", "5–12×", "25–45×", "n/a", "SaaS-multiple lens. Premium for high-recurring revenue and operating leverage."),
        ("Wealth-platform / financial-platform fintechs", "Plaid, Stripe (private), Adyen", "10–20×", "30–60×", "n/a", "Aspirational. Only relevant if FP can credibly position as a SaaS infrastructure play."),
        ("Implied FP blended (mid-range)", "—", "5–8×", "10–15×", "10–30 bps", "Suggested blended target reflecting hybrid model. Validate against deal flow."),
    ]

    for cat, names, evrev, evebitda, evaum, notes in comps:
        ws.cell(row=row, column=1, value=cat).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=names).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=evrev).font = NORMAL_FONT
        ws.cell(row=row, column=4, value=evebitda).font = NORMAL_FONT
        ws.cell(row=row, column=5, value=evaum).font = NORMAL_FONT
        ws.cell(row=row, column=6, value=notes).font = NORMAL_FONT
        ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True)
        if "Implied FP" in cat:
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = TOTAL_FILL
                ws.cell(row=row, column=col).font = TOTAL_FONT
        row += 1



def build_metrics_dashboard(wb, pnl_likely: dict, metrics: dict):
    ws = wb.create_sheet("Metrics_Dashboard")
    ws.column_dimensions["A"].width = 48
    for c in "BCDEFG":
        ws.column_dimensions[c].width = 14
    ws.column_dimensions["H"].width = 50

    style_title_row(ws, 1, "VC Metrics Dashboard — Likely Case (USD)", span=8)
    ws.cell(row=2, column=1, value=(
        "Operational and capital-efficiency metrics typically asked for at Series A diligence. "
        "All from the Likely-case P&L. Briefing-deck slide 13 cited LTV:CAC = 65:1; this tab traces it from the model."
    )).font = NOTE_FONT
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    ws.row_dimensions[2].height = 35

    row = 4
    headers = ["Metric"] + YEARS + ["Notes"]
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=i + 1, value=h)
        c.font = H2_FONT
        c.fill = LIGHT_FILL
    row += 1

    def write_metric_row(label, vals, fmt=USD_FMT, note="", bold=False, fill=None):
        nonlocal row
        c = ws.cell(row=row, column=1, value=label)
        c.font = TOTAL_FONT if bold else NORMAL_FONT
        if fill:
            c.fill = fill
        for i, v in enumerate(vals):
            cell = ws.cell(row=row, column=2 + i, value=v)
            cell.font = TOTAL_FONT if bold else NORMAL_FONT
            cell.number_format = fmt
            if fill:
                cell.fill = fill
        c2 = ws.cell(row=row, column=8, value=note)
        c2.font = NOTE_FONT
        c2.alignment = Alignment(wrap_text=True)
        row += 1

    style_section_row(ws, row, "Operational Efficiency", span=8); row += 1
    write_metric_row("Revenue per FTE (recurring)", metrics["rev_per_fte"], fmt=USD_FMT, note="At Y3+ this is best-in-class for regulated FS.")
    write_metric_row("AUM per FTE (USD)", metrics["aum_per_fte"], fmt=USD_M_FMT, note="Benchmark: specialty insurer ~$200–500M / FTE.")
    write_metric_row("Revenue per B2B licensee", metrics["rev_per_licensee"], fmt=USD_M_FMT, note="Per cumulative Product Issuer.")

    style_section_row(ws, row, "Growth + Profitability", span=8); row += 1
    write_metric_row("Recurring revenue growth (YoY)", metrics["rev_growth"], fmt=PCT_FMT, note="Y2 growth jumps from low base; normalises Y3+.")
    write_metric_row("EBITDA margin (recurring)", metrics["ebitda_margin_recurring"], fmt=PCT_FMT, note="Recurring-only margin (ex profit-share).")
    write_metric_row("Rule of 40 (Growth + Margin)", metrics["rule_of_40"], fmt=PCT_FMT, note=">40% = top-quartile SaaS efficiency.", bold=True, fill=EBITDA_FILL)

    style_section_row(ws, row, "Customer Economics (B2B Product Issuers)", span=8); row += 1

    ws.cell(row=row, column=1, value="LTV per B2B customer (5-yr cumulative revenue, undiscounted)").font = NORMAL_FONT
    cell = ws.cell(row=row, column=2, value=metrics["ltv_per_customer_5yr"])
    cell.number_format = USD_M_FMT
    cell.font = TOTAL_FONT
    ws.cell(row=row, column=8, value="Total Y1-Y5 revenue / Y5 cumulative licensees.").font = NOTE_FONT
    row += 1

    ws.cell(row=row, column=1, value="CAC per B2B customer (sales + legal + onboarding)").font = NORMAL_FONT
    cell = ws.cell(row=row, column=2, value=metrics["cac_per_customer"])
    cell.number_format = USD_FMT
    ws.cell(row=row, column=8, value="From Drivers tab + briefing-deck slide 13.").font = NOTE_FONT
    row += 1

    ws.cell(row=row, column=1, value="LTV : CAC ratio").font = TOTAL_FONT
    cell = ws.cell(row=row, column=2, value=metrics["ltv_cac_ratio"])
    cell.number_format = "0.0\"×\""
    cell.font = TOTAL_FONT
    cell.fill = EBITDA_FILL
    ws.cell(row=row, column=8, value="Briefing deck cited 65:1. SaaS norm is 3–5:1.").font = NOTE_FONT
    row += 1

    ws.cell(row=row, column=1, value="Cash payback period (years)").font = NORMAL_FONT
    cell = ws.cell(row=row, column=2, value=metrics["cash_payback_years"])
    cell.number_format = "0.00\" yrs\""
    ws.cell(row=row, column=8, value="CAC / per-customer Y3 revenue. <1 yr is exceptional.").font = NOTE_FONT
    row += 2

    ws.cell(row=row, column=1, value=(
        "These metrics flow into the deck and FAQ. The Rule-of-40 figure is the SaaS-comparable growth-quality test; "
        "the LTV:CAC ratio is the capital-efficiency test that VCs use to size follow-on investment thesis."
    )).font = NOTE_FONT
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    # Compute results
    results = {case.name: compute_pnl(case) for case in CASES}

    # Validate Likely case approximates v10.1
    likely_y3_rev = results["Likely"]["rev_total"][3]
    likely_y5_rev = results["Likely"]["rev_total"][5]
    print(f"Likely Y3 total revenue: ${likely_y3_rev/1e6:.1f}M (v10.1: $81.8M)")
    print(f"Likely Y5 total revenue: ${likely_y5_rev/1e6:.1f}M (v10.1: $930.8M)")
    print(f"Likely Y3 EBITDA margin (recurring): {results['Likely']['ebitda_pct_recurring'][3]:.1%}")
    print(f"Conservative Y5 total revenue: ${results['Conservative']['rev_total'][5]/1e6:.1f}M")
    print(f"Optimistic Y5 total revenue: ${results['Optimistic']['rev_total'][5]/1e6:.1f}M")

    # Build workbook
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet, README will be inserted at index 0

    # Compute extrapolation + valuation + metrics for all 3 cases
    extrap_by_case = {c.name: compute_extrapolation(c, results[c.name]) for c in CASES}
    dcf_by_case = {
        c.name: compute_dcf(results[c.name], extrap_by_case[c.name], wacc=0.14, terminal_growth=0.03, tax_rate=0.30)
        for c in CASES
    }
    vc_by_case = {c.name: compute_vc_method(results[c.name], extrap_by_case[c.name]) for c in CASES}

    likely = results["Likely"]
    metrics = compute_metrics(likely, extrap_by_case["Likely"])

    print(f"\n--- DCF Enterprise Value by case (WACC 14%, terminal g 3%) ---")
    for c in CASES:
        ev = dcf_by_case[c.name]["enterprise_value"]
        y10_ebitda = extrap_by_case[c.name]["normalised_ebitda"][-1]
        print(f"  {c.name:13} Y10 EBITDA ${y10_ebitda/1e6:.0f}M  →  EV ${ev/1e9:.2f}B")
    print(f"  Likely LTV:CAC: {metrics['ltv_cac_ratio']:.1f}×")

    build_readme(wb)
    build_drivers(wb)
    for case in CASES:
        build_pnl(wb, case, results[case.name])
    build_summary(wb, results)
    build_au_only(wb, results)
    build_surplus_sensitivity(wb, results)
    build_runway(wb, results)
    build_ai_cost_detail(wb, results)
    build_extrapolation(wb, results, extrap_by_case)
    build_valuation(wb, results, extrap_by_case, dcf_by_case, vc_by_case)
    build_metrics_dashboard(wb, likely, metrics)

    wb.save(OUT)
    print(f"\nSaved: {OUT}")


if __name__ == "__main__":
    main()
