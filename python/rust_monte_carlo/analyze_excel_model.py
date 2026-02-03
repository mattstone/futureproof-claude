#!/usr/bin/env python3
"""
Analyze Excel model and compare with Python implementation
"""

import openpyxl
import pandas as pd
import sys

excel_file = "../Copy of 20250624 FutureProof Mini Model Data Room.xlsm"

print("=" * 90)
print("EXCEL MODEL ANALYSIS")
print("=" * 90)

# Load workbook
wb = openpyxl.load_workbook(excel_file, data_only=False, keep_vba=True)

print(f"\n📊 Workbook: {excel_file}")
print(f"Sheets found: {len(wb.sheetnames)}")
print(f"Sheet names: {wb.sheetnames}")

# Analyze each sheet
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\n{'=' * 90}")
    print(f"SHEET: {sheet_name}")
    print(f"{'=' * 90}")

    # Get dimensions
    max_row = sheet.max_row
    max_col = sheet.max_column
    print(f"Dimensions: {max_row} rows × {max_col} columns")

    # Find named ranges or key cells
    print(f"\nSearching for key parameters and calculations...")

    key_terms = [
        'loan', 'lvr', 'house', 'value', 'duration', 'annuity',
        'interest', 'rate', 'cash', 'margin', 'wholesale',
        'insurance', 'cost', 'hedg', 'holiday', 'enter', 'exit',
        'equity', 'return', 'volatility', 'monte', 'carlo',
        'reinvest', 'deficit', 'xirr', 'cagr', 'npv'
    ]

    found_cells = []

    # Scan first 50 rows and 20 columns for key terms
    for row in range(1, min(51, max_row + 1)):
        for col in range(1, min(21, max_col + 1)):
            cell = sheet.cell(row, col)
            if cell.value:
                cell_str = str(cell.value).lower()
                for term in key_terms:
                    if term in cell_str:
                        # Get the value in the next cell (often the parameter value)
                        next_cell = sheet.cell(row, col + 1)
                        found_cells.append({
                            'location': f"{cell.coordinate}",
                            'label': cell.value,
                            'value': next_cell.value if next_cell.value else "(formula/empty)"
                        })
                        break

    if found_cells:
        print(f"\n📋 Found {len(found_cells)} potential parameter cells:")
        for item in found_cells[:30]:  # Limit to first 30
            print(f"  {item['location']}: {item['label']} = {item['value']}")

    # Look for formulas
    print(f"\n🔍 Scanning for complex formulas...")
    formula_cells = []

    for row in range(1, min(201, max_row + 1)):
        for col in range(1, min(21, max_col + 1)):
            cell = sheet.cell(row, col)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formula = cell.value
                # Look for interesting formulas (IF, SUM, NPV, XIRR, etc.)
                if any(func in formula.upper() for func in ['IF(', 'SUM(', 'NPV(', 'XIRR(', 'IRR(', 'MAX(', 'MIN(', 'AVERAGE(']):
                    formula_cells.append({
                        'location': cell.coordinate,
                        'formula': formula[:100] + '...' if len(formula) > 100 else formula
                    })

    if formula_cells:
        print(f"\n📐 Found {len(formula_cells)} complex formulas (showing first 20):")
        for item in formula_cells[:20]:
            print(f"  {item['location']}: {item['formula']}")

# Check for VBA macros
print(f"\n{'=' * 90}")
print("VBA MACROS")
print(f"{'=' * 90}")

if hasattr(wb, 'vba_archive') and wb.vba_archive:
    print("\n✅ VBA macros found in workbook")
    print("Note: Full VBA extraction requires additional processing")
    print("Common macro functions in Excel financial models:")
    print("  - Monte Carlo simulation loops")
    print("  - Custom XIRR/NPV calculations")
    print("  - Path generation (GBM)")
    print("  - Scenario analysis")
else:
    print("\n❌ No VBA macros detected (or unable to extract)")

print(f"\n{'=' * 90}")
print("NEXT STEPS FOR DETAILED COMPARISON")
print(f"{'=' * 90}")
print("""
1. Extract specific parameter values from Excel sheets
2. Compare Excel formulas with Python calculations
3. Check for differences in:
   - Interest rate calculation (quarterly vs monthly)
   - Payment holiday logic (entry/exit conditions)
   - Reinvestment calculations
   - Insurance cost timing
   - Hedging implementation
   - XIRR/NPV calculations
4. If VBA macros exist, extract and compare with Python simulation logic
""")

# Try to export key sheets to CSV for manual inspection
print(f"\n📤 Exporting sheets to CSV for inspection...")

for sheet_name in wb.sheetnames[:3]:  # Export first 3 sheets
    try:
        sheet = wb[sheet_name]
        data = []
        for row in sheet.iter_rows(min_row=1, max_row=min(200, sheet.max_row),
                                    min_col=1, max_col=min(20, sheet.max_column)):
            row_data = []
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    row_data.append(f"FORMULA: {cell.value[:50]}")
                else:
                    row_data.append(cell.value)
            data.append(row_data)

        df = pd.DataFrame(data)
        output_file = f"excel_export_{sheet_name.replace(' ', '_')}.csv"
        df.to_csv(output_file, index=False, header=False)
        print(f"  ✅ Exported {sheet_name} to {output_file}")
    except Exception as e:
        print(f"  ❌ Failed to export {sheet_name}: {e}")

print("\n" + "=" * 90)
